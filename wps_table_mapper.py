#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
WPS表格字段映射工具（麒麟Linux版）
=================================

单文件 PyQt5 桌面应用，用于把「WPS 源表格」的字段按自定义规则映射并写入「WPS 目标表格」。

功能要点
--------
1. 文件上传：分别加载源表(.xlsx/.xls) 与目标表(.xlsx/.xls)；
            源表读取表头 + 全部数据；目标表仅读取表头（第一行），不读数据。
2. 映射配置：QTableWidget 默认 20 行，每行含
            填入方式(直接复制/合并复制/拆分复制)、源列、目标列、连接符、拆分数、操作(+/-)。
3. 数据处理：日期单元格自动识别并转换为 "YYYY.MM"；
            直接复制 / 合并复制(2~5 源列 + 连接符) / 拆分复制(把源列值复制到所选 N 个目标列)。
4. 执行保存：写入目标表指定列（保留原格式/样式/公式，仅覆盖数据区），另存为新 .xlsx。
5. 交互：状态栏提示、非法操作弹窗警告。

依赖
----
    # 通用（Windows/macOS/x86 Linux）：
    pip install PyQt5 openpyxl
    # 读取 .xls 需要额外安装：
    pip install xlrd

    # 麒麟 Linux / ARM64 / 龙芯 等国产平台（优先 apt，避免 pip 没有预编译包）：
    sudo apt update
    sudo apt install -y python3-pyqt5 python3-openpyxl
    # 如需读取 .xls 再加装：
    sudo apt install -y python3-xlrd

    # 【关键】麒麟源常见报错处理：
    # 若提示「python3-pyqt5.qtsvg : 依赖: python3-pyqt5 (= 5.14.1...) 但是 5.15.4... 正要被安装」：
    #   → 去掉 qtsvg（代码未用到），冲突部分用 pip 补齐：
    #     sudo apt install -y python3-pyqt5
    #     python3 -m pip install openpyxl xlrd

说明
----
- openpyxl 原生只支持 .xlsx/.xlsm。对 .xls 文件采用 xlrd 读取（仅取表头/数据），
  且 .xls 目标表无法保留原格式，输出会生成新的 .xlsx（程序会给出提示）。
- 麒麟 Linux 通过 apt 安装 python3-pyqt5 即可满足运行；不要装 python3-pyqt5.qtsvg
  （本程序未使用 QtSvg，且麒麟源内该包常与新版 python3-pyqt5 产生版本冲突）。
- 程序导入时 PyQt5 失败会自动回退尝试 PyQt6；两种绑定都能正常运行。
"""

import sys
import os
import re
import json
import datetime

# ---------- 平台与架构检测 ----------
def _detect_platform():
    """返回 (os_family, arch, is_kylin_linux)。"""
    import platform
    os_name = platform.system().lower()
    arch = platform.machine().lower()
    is_kylin = False
    if os_name == "linux":
        try:
            with open("/etc/os-release", "r", encoding="utf-8") as f:
                content = f.read().lower()
                if "kylin" in content or "银河麒麟" in content:
                    is_kylin = True
        except Exception:
            pass
    return os_name, arch, is_kylin


# ---------- 依赖检查 ----------
_PYQT_VER = None   # 运行时绑定："PyQt5" 或 "PyQt6"

def _install_hint_pyqt(detail_err=""):
    """根据当前平台生成最合适的 PyQt 安装指引。"""
    os_name, arch, is_kylin = _detect_platform()
    lines = []
    lines.append(f"无法加载 PyQt5/PyQt6（{detail_err}）")
    lines.append(f"当前解释器：{sys.executable}")
    lines.append(f"平台信息：os={os_name} arch={arch}")
    lines.append("")

    if is_kylin or (os_name == "linux" and arch in ("aarch64", "arm64", "loongarch64", "sw_64")):
        # 麒麟 / ARM64 / 龙芯 / 申威 等国产平台：优先 apt 系统包
        lines.append("【推荐】使用系统包管理器安装（ARM/国产平台通常没有 PyPI 预编译包）：")
        lines.append("  sudo apt update")
        lines.append("  # 注：不装 python3-pyqt5.qtsvg，避免麒麟源内 pyqt5 版本冲突（代码未使用 QtSvg）")
        lines.append("  sudo apt install -y python3-pyqt5 python3-openpyxl")
        lines.append("  # 如需读取 .xls 再加装：sudo apt install -y python3-xlrd")
        lines.append("")
        lines.append("【麒麟源版本冲突处理】若出现「依赖: python3-pyqt5 (= 5.14.1...) 但是 5.15.4... 正要被安装」这类错误：")
        lines.append("  方案 A（推荐）——只装能装上的，剩余用 pip：")
        lines.append("    sudo apt install -y python3-pyqt5")
        lines.append(f"    {sys.executable} -m pip install openpyxl xlrd")
        lines.append("  方案 B——用 aptitude 自动降级解决冲突（需要先装 aptitude）：")
        lines.append("    sudo apt install -y aptitude && sudo aptitude install -y python3-pyqt5 python3-openpyxl")
        lines.append("")
        lines.append("【备选】如系统源完全无可用版本，再试 pip（需要有编译环境，成功率低）：")
        lines.append(f"  {sys.executable} -m pip install --upgrade pip")
        lines.append(f"  {sys.executable} -m pip install PyQt5 openpyxl xlrd")
    elif os_name == "linux":
        lines.append("【推荐】使用系统包管理器：")
        lines.append("  sudo apt install -y python3-pyqt5 python3-openpyxl")
        lines.append("  # 麒麟源若遇到 qt.svg 版本冲突，只装 python3-pyqt5，剩余用 pip 补：")
        lines.append(f"  #   {sys.executable} -m pip install openpyxl xlrd")
        lines.append("")
        lines.append("或使用 pip：")
        lines.append(f"  {sys.executable} -m pip install PyQt5 openpyxl xlrd")
    elif os_name == "darwin":
        lines.append(f"请运行：")
        lines.append(f"  {sys.executable} -m pip install PyQt5 openpyxl xlrd")
    else:
        # Windows 等
        lines.append(f"请运行（需要与当前解释器一致）：")
        lines.append(f"  {sys.executable} -m pip install --upgrade pip")
        lines.append(f"  {sys.executable} -m pip install PyQt5 openpyxl xlrd")
    return "\n".join(lines) + "\n"


# 先探测 sip 情况，用于区分是「模块缺失」还是「sip 版本不匹配」
try:
    import sip as _sip_mod
    _sip_info = f"sip={_sip_mod.__file__}"
except Exception as _e:
    _sip_info = f"sip不可用: {_e}"

# 1) 优先尝试 PyQt5
_import_err = ""
try:
    from PyQt5.QtWidgets import (
        QApplication, QMainWindow, QWidget, QVBoxLayout, QHBoxLayout,
        QPushButton, QLabel, QComboBox, QLineEdit, QSpinBox, QTableWidget,
        QFileDialog, QMessageBox, QHeaderView, QDialog, QListWidget,
        QListWidgetItem, QGroupBox, QAbstractItemView, QTabWidget,
        QRadioButton, QFormLayout, QTableWidgetItem, QGridLayout,
    )
    from PyQt5.QtCore import Qt, QSize, pyqtSignal, QTimer
    from PyQt5.QtGui import QFont, QColor, QIcon, QPixmap
    _PYQT_VER = "PyQt5"
    # PyQt5 使用 exec_()
    def _app_exec(app):
        return app.exec_()
except ImportError as _e1:
    _import_err = f"PyQt5: {_e1} | {_sip_info}"
    # 2) 回退 PyQt6（部分新系统只有 PyQt6）
    try:
        from PyQt6.QtWidgets import (
            QApplication, QMainWindow, QWidget, QVBoxLayout, QHBoxLayout,
            QPushButton, QLabel, QComboBox, QLineEdit, QSpinBox, QTableWidget,
            QFileDialog, QMessageBox, QHeaderView, QDialog, QListWidget,
            QListWidgetItem, QGroupBox, QAbstractItemView, QTabWidget,
            QRadioButton, QFormLayout, QTableWidgetItem, QGridLayout,
        )
        from PyQt6.QtCore import Qt, QSize, pyqtSignal, QTimer
        from PyQt6.QtGui import QFont, QColor, QIcon, QPixmap
        _PYQT_VER = "PyQt6"
        # PyQt6 与 PyQt5 的兼容差异适配：exec_ → exec
        def _app_exec(app):
            return app.exec() if hasattr(app, "exec") else app.exec_()
    except ImportError as _e2:
        _import_err += f"; PyQt6: {_e2}"
        sys.stderr.write(_install_hint_pyqt(_import_err))
        sys.exit(1)

try:
    import openpyxl
except ImportError as e:
    os_name, arch, is_kylin = _detect_platform()
    hint_lines = [f"缺少 openpyxl（{e}）"]
    if is_kylin or (os_name == "linux" and arch in ("aarch64", "arm64", "loongarch64", "sw_64")):
        hint_lines.append("【推荐】使用系统包管理器：")
        hint_lines.append("  sudo apt install -y python3-openpyxl")
        hint_lines.append("或使用 pip：")
    hint_lines.append(f"  {sys.executable} -m pip install openpyxl")
    sys.stderr.write("\n".join(hint_lines) + "\n")
    sys.exit(1)


# ---------- PyQt5 / PyQt6 枚举兼容层 ----------
# PyQt6 枚举都在各自的命名空间下，这里把 PyQt5 风格的扁平引用补到全局
if _PYQT_VER == "PyQt6":
    # Qt 命名空间
    from PyQt6.QtCore import Qt as _QtNS
    for _name in dir(_QtNS):
        _obj = getattr(_QtNS, _name)
        # 对枚举类，把其成员值挂到 Qt 顶层，模拟 PyQt5 行为
        if isinstance(_obj, type) and issubclass(_obj, int):
            for _mname, _mval in _obj.__members__.items():
                if not hasattr(Qt, _mname):
                    setattr(Qt, _mname, _mval)
    # QDialog::DialogCode
    for _mname, _mval in QDialog.DialogCode.__members__.items():
        if not hasattr(QDialog, _mname):
            setattr(QDialog, _mname, _mval)
    # QHeaderView::ResizeMode
    for _mname, _mval in QHeaderView.ResizeMode.__members__.items():
        if not hasattr(QHeaderView, _mname):
            setattr(QHeaderView, _mname, _mval)
    # QAbstractItemView::SelectionMode / SelectionBehavior / EditTrigger
    for _cls_attr in ("SelectionMode", "SelectionBehavior", "EditTrigger"):
        if hasattr(QAbstractItemView, _cls_attr):
            _ns = getattr(QAbstractItemView, _cls_attr)
            if isinstance(_ns, type):
                for _mname, _mval in _ns.__members__.items():
                    if not hasattr(QAbstractItemView, _mname):
                        setattr(QAbstractItemView, _mname, _mval)
    # QTabWidget / QTabBar 相关
    for _cls, _attrs in (
        (QTabWidget, ("TabPosition",)),
        (QTabBar if "QTabBar" in globals() else None, ()),
    ):
        if _cls is None:
            continue
        for _cls_attr in _attrs:
            if hasattr(_cls, _cls_attr):
                _ns = getattr(_cls, _cls_attr)
                if isinstance(_ns, type):
                    for _mname, _mval in _ns.__members__.items():
                        if not hasattr(_cls, _mname):
                            setattr(_cls, _mname, _mval)


# ---------- 映射表格列定义 ----------
COL_SRCTBL = 0  # 源表选择（多源表格时选择数据来自哪个源表）
COL_METHOD = 1  # 填入方式
COL_SOURCE = 2  # 源列
COL_TARGET = 3  # 目标列
COL_SEP    = 4  # 连接符（合并复制用）
COL_SPLIT  = 5  # 拆分数（拆分复制用）
COL_ACTION = 6  # 操作（+ / -）
MAP_COL_COUNT = 7  # 映射表总列数


# ---------- 可拖拽行排序的表格 ----------
class MappingTable(QTableWidget):
    """通过左侧行号表头拖拽整行重排序；点击单元格内的 QComboBox 任意区域即可展开下拉。"""
    rowReorderRequested = pyqtSignal(list)  # 新的行顺序（逻辑行索引列表）

    def __init__(self, parent=None):
        super().__init__(parent)
        self.setSelectionMode(QAbstractItemView.SingleSelection)
        self.setSelectionBehavior(QAbstractItemView.SelectRows)
        vh = self.verticalHeader()
        vh.setSectionsMovable(True)     # 行号可拖动重排
        vh.setDragEnabled(True)
        vh.setMinimumWidth(40)          # 加宽行号，方便抓取
        vh.sectionMoved.connect(self._on_section_moved)

    def mousePressEvent(self, event):
        """点击单元格内 QComboBox 的任意区域（包含文本部分）都直接展开下拉。"""
        if event.button() == Qt.LeftButton:
            pos = event.pos()
            idx = self.indexAt(pos)
            if idx.isValid():
                w = self.cellWidget(idx.row(), idx.column())
                if isinstance(w, QComboBox):
                    # 稍微延迟触发，避免默认行为覆盖
                    QTimer.singleShot(0, w.showPopup)
                    event.accept()
                    return
        super().mousePressEvent(event)

    def _on_section_moved(self, logicalIndex, oldVisualIndex, newVisualIndex):
        n = self.rowCount()
        if n <= 1:
            return
        hdr = self.verticalHeader()
        order = [hdr.logicalIndex(p) for p in range(n)]
        self.rowReorderRequested.emit(order)


# ---------- 日期识别与转换 ----------
DATE_FORMATS = [
    "%Y-%m-%d", "%Y/%m/%d", "%Y.%m.%d", "%Y年%m月%d日",
    "%Y-%m-%d %H:%M:%S", "%Y/%m/%d %H:%M:%S",
    "%Y-%m", "%Y/%m", "%Y.%m", "%Y%m%d",
]


def is_date_value(v):
    """判断值是否为日期（datetime 对象或可解析为日期的字符串）。"""
    if isinstance(v, (datetime.datetime, datetime.date)):
        return True
    if isinstance(v, str):
        s = v.strip()
        for fmt in DATE_FORMATS:
            try:
                datetime.datetime.strptime(s, fmt)
                return True
            except ValueError:
                pass
    return False


def format_date(v):
    """把日期值格式化为 'YYYY.MM'。"""
    if isinstance(v, datetime.datetime):
        return f"{v.year}.{v.month:02d}"
    if isinstance(v, datetime.date):
        return f"{v.year}.{v.month:02d}"
    s = str(v).strip()
    for fmt in DATE_FORMATS:
        try:
            d = datetime.datetime.strptime(s, fmt)
            return f"{d.year}.{d.month:02d}"
        except ValueError:
            pass
    return s


def normalize(v):
    """数据写入前的标准化：日期 -> 'YYYY.MM'，其余原样返回（数字保持数字）。"""
    if v is None:
        return None
    if is_date_value(v):
        return format_date(v)
    return v


def _sort_key(v):
    """排序键：按类型分桶（None < 数字 < 日期 < 字符串/其它），桶内按值比较。

    这样混合类型（如数字与字符串混在一列）不会在比较时抛 TypeError，
    且只决定顺序、不修改原始值，满足「保持所有原始数据类型不变」。
    """
    if v is None:
        return (0, 0, "")
    if isinstance(v, bool):
        return (3, 0, str(v))
    if isinstance(v, (int, float)):
        return (1, v, "")
    if isinstance(v, (datetime.datetime, datetime.date)):
        return (2, 0, v.isoformat())
    return (3, 0, str(v))


def _tab_icon(color):
    """生成一个纯色小方块图标，用于在标签页上做颜色区分（页面本身不上色）。"""
    px = QPixmap(16, 16)
    px.fill(QColor(color))
    return QIcon(px)


# ---------- 多选对话框（合并复制选择多个源列） ----------
class MultiSelectDialog(QDialog):
    def __init__(self, items, checked, parent=None, max_sel=None, title="选择列"):
        super().__init__(parent)
        self.max_sel = max_sel if (max_sel and max_sel > 0) else len(items)
        self.setWindowTitle(title)
        self.resize(320, 380)
        layout = QVBoxLayout(self)

        self.listw = QListWidget()
        for it in items:
            item = QListWidgetItem(it)
            item.setFlags(item.flags() | Qt.ItemIsUserCheckable)
            item.setCheckState(Qt.Checked if it in (checked or []) else Qt.Unchecked)
            self.listw.addItem(item)
        self.listw.itemChanged.connect(self._on_item_changed)
        layout.addWidget(self.listw)

        layout.addWidget(QLabel(f"提示：最多可选择 {self.max_sel} 个列，超过自动取消勾选。"))

        btns = QHBoxLayout()
        ok = QPushButton("确定")
        cancel = QPushButton("取消")
        ok.clicked.connect(self.accept)
        cancel.clicked.connect(self.reject)
        btns.addWidget(ok)
        btns.addWidget(cancel)
        layout.addLayout(btns)

    def _on_item_changed(self, item):
        if len(self.selected()) > self.max_sel:
            item.setCheckState(Qt.Unchecked)

    def selected(self):
        res = []
        for i in range(self.listw.count()):
            it = self.listw.item(i)
            if it.checkState() == Qt.Checked:
                res.append(it.text())
        return res[:self.max_sel]


# ---------- 主窗口 ----------
class MainWindow(QMainWindow):
    def __init__(self):
        super().__init__()
        self.source_path = ""
        self.target_path = ""
        self.source_headers = []
        self.target_headers = []
        self.target_col_count = 0
        self.source_header_row = 1   # 表头所在行（1-based），兼容合并标题行
        self.target_header_row = 1
        self.source_data = []   # list[list]，不含表头
        self.source_sheet_name = ""   # 当前源表子表名（主源表，向后兼容）
        self.target_sheet_name = ""   # 当前目标表子表名

        # ---- 多源表格管理 ----
        # self.sources: list[dict]，每项 = {path, sheet, headers, data, header_row, label}
        # self.source_widgets: list[dict]，每项 = {row_widget, btn_load, lbl_name, combo_sheet, btn_del}
        self.sources = []
        self.source_widgets = []
        self.MAX_SOURCES = 5

        # 表格合并标签页状态
        self.merge_files = []            # 已添加的文件路径列表
        self.merge_sheets = []           # [{path, headers, data}, ...]
        self.merge_canonical_headers = []  # 合并后使用的规范表头（取首个文件）
        self.merge_title_rows = []        # 表头上方内容（如合并标题/表名），取首个文件
        self.merge_result = None         # (headers, rows) 合并排序后的结果
        self._merge_preview_spans = []   # 预览表格中已设置的跨列合并记录

        self.setWindowTitle("WPS表格字段映射工具（麒麟Linux版）")
        # 启动时按屏幕尺寸自适应放大（避免界面偏小）
        screen = QApplication.primaryScreen()
        sg = screen.availableGeometry() if screen else None
        if sg:
            w = min(max(int(sg.width() * 0.85), 1200), 1600)
            h = min(max(int(sg.height() * 0.88), 820), 1000)
        else:
            w, h = 1440, 920
        self.resize(w, h)
        self.setMinimumSize(1200, 820)

        self._build_ui()
        self._init_mapping(10)
        self.statusBar().showMessage("就绪：请先上传源表格与目标表格")

    # ========== UI 构建 ==========
    def _on_tab_changed(self, index):
        """切换标签时清理底部状态栏：合并标签有自己的状态行，不应显示映射相关的提示。"""
        # 索引 1 = 「表格合并」标签，索引 0 = 「表格映射」标签
        if index == 1:
            self.statusBar().clearMessage()
        elif index == 0:
            self.statusBar().showMessage("就绪：请先上传源表格与目标表格")

    def _build_ui(self):
        central = QWidget()
        self.setCentralWidget(central)
        outer = QVBoxLayout(central)
        outer.setSpacing(10)
        outer.setContentsMargins(12, 12, 12, 12)

        tabs = QTabWidget()
        self.tabs = tabs

        # ---- 表格映射 标签页 ----
        mapping_widget = QWidget()
        root = QVBoxLayout(mapping_widget)
        root.setSpacing(10)
        root.setContentsMargins(12, 12, 12, 12)

        # ---- 顶部行：文件上传区 + 操作按钮 并排 ----
        top_row = QHBoxLayout()
        top_row.setSpacing(16)

        # ---- 文件上传区 ----
        up_group = QGroupBox("文件上传")
        up_layout = QVBoxLayout(up_group)
        up_layout.setSpacing(8)
        up_layout.setContentsMargins(10, 12, 10, 10)

        # ---- 源表格管理区（支持多个源表，最多5个）----
        src_label_row = QHBoxLayout()
        src_lbl = QLabel("源表格（最多 5 个）：")
        # 固定宽度：上传按钮(120) + 间距(8) + 文件名(340) + 间距(8)，使"添加源表格"与下方"子表:"左对齐
        src_lbl.setFixedWidth(640)
        src_label_row.addWidget(src_lbl)
        self.btn_add_source = QPushButton("+ 添加源表格")
        self.btn_add_source.setMinimumWidth(120)
        self.btn_add_source.setMinimumHeight(32)
        self.btn_add_source.setCursor(Qt.PointingHandCursor)
        self.btn_add_source.setStyleSheet("""
            QPushButton {
                background-color: #eff6ff;
                border: 1px solid #2563eb;
                border-radius: 6px;
                color: #1e40af;
                font-weight: bold;
                padding: 4px 12px;
                text-align: center;
            }
            QPushButton:hover {
                background-color: #dbeafe;
                border: 1px solid #1d4ed8;
            }
            QPushButton:pressed {
                background-color: #bfdbfe;
            }
            QPushButton:disabled {
                background-color: #f3f4f6;
                border: 1px solid #d1d5db;
                color: #9ca3af;
            }
        """)
        self.btn_add_source.clicked.connect(self._add_source_row)
        src_label_row.addWidget(self.btn_add_source)
        src_label_row.addStretch(1)
        up_layout.addLayout(src_label_row)

        # 源表行容器
        self.src_container = QVBoxLayout()
        self.src_container.setSpacing(4)
        up_layout.addLayout(self.src_container)

        # 初始创建第一个源表行（主源表，不可删除）
        self._add_source_row(primary=True)

        # ---- 目标表格上传行 ----
        tgt_row = QHBoxLayout()
        tgt_row.setSpacing(8)
        self.btn_load_target = QPushButton("上传目标表格")
        self.btn_load_target.setMinimumWidth(120)
        self.btn_load_target.setMinimumHeight(36)
        self.btn_load_target.setStyleSheet("""
            QPushButton {
                background-color: #fef3c7;
                border: 1px solid #d97706;
                border-radius: 6px;
                color: #92400e;
                font-weight: bold;
                padding: 4px 14px;
                text-align: center;
            }
            QPushButton:hover {
                background-color: #fde68a;
                border: 1px solid #b45309;
            }
            QPushButton:pressed {
                background-color: #fcd34d;
            }
        """)
        self.btn_load_target.clicked.connect(self.load_target_file)
        self.lbl_target = QLabel("（未选择目标文件）")
        self.lbl_target.setWordWrap(False)
        self.lbl_target.setFixedWidth(440)
        tgt_row.addWidget(self.btn_load_target)
        tgt_row.addWidget(self.lbl_target)
        tgt_row.addWidget(QLabel("子表:"))
        self.combo_target_sheet = QComboBox()
        self.combo_target_sheet.setEnabled(False)
        self.combo_target_sheet.setMinimumWidth(200)   # 10 个汉字宽度
        self.combo_target_sheet.setToolTip("选择目标表中的子表（Sheet）")
        self.combo_target_sheet.currentIndexChanged.connect(self._on_target_sheet_changed)
        tgt_row.addWidget(self.combo_target_sheet)
        tgt_row.addSpacing(12)
        tgt_row.addWidget(QLabel("表头行:"))
        self.spin_target_header = QSpinBox()
        self.spin_target_header.setMinimum(1)
        self.spin_target_header.setMaximum(50)
        self.spin_target_header.setValue(1)
        self.spin_target_header.setAlignment(Qt.AlignCenter)
        self.spin_target_header.setMinimumWidth(90)
        self.spin_target_header.setPrefix("第 ")
        self.spin_target_header.setSuffix(" 行")
        self.spin_target_header.valueChanged.connect(
            lambda: self._apply_target_load() if getattr(self, "target_path", None) else None
        )
        tgt_row.addWidget(self.spin_target_header)
        tgt_row.addStretch(1)
        up_layout.addLayout(tgt_row)

        top_row.addWidget(up_group, 7)  # 文件上传区占 70%

        # ---- 操作按钮区（右侧，2 列网格）----
        exec_panel = QGroupBox("操作")
        exec_layout = QGridLayout(exec_panel)
        exec_layout.setHorizontalSpacing(12)
        exec_layout.setVerticalSpacing(12)
        exec_layout.setContentsMargins(14, 18, 14, 14)

        # 导出配置：橙色主题
        self.btn_export = QPushButton("导出配置")
        self.btn_export.setMinimumHeight(44)
        self.btn_export.setStyleSheet("""
            QPushButton {
                background-color: #fff7ed;
                border: 1px solid #ea580c;
                border-radius: 8px;
                color: #9a3412;
                font-weight: bold;
                padding: 6px 14px;
            }
            QPushButton:hover {
                background-color: #ffedd5;
                border: 1px solid #c2410c;
            }
            QPushButton:pressed {
                background-color: #fed7aa;
            }
            QPushButton:disabled {
                background-color: #f3f4f6;
                border: 1px solid #d1d5db;
                color: #9ca3af;
            }
        """)
        self.btn_export.clicked.connect(self.export_strategy)
        exec_layout.addWidget(self.btn_export, 0, 0)

        # 导入配置：紫色主题
        self.btn_import = QPushButton("导入配置")
        self.btn_import.setMinimumHeight(44)
        self.btn_import.setStyleSheet("""
            QPushButton {
                background-color: #f5f3ff;
                border: 1px solid #8b5cf6;
                border-radius: 8px;
                color: #5b21b6;
                font-weight: bold;
                padding: 6px 14px;
            }
            QPushButton:hover {
                background-color: #ede9fe;
                border: 1px solid #7c3aed;
            }
            QPushButton:pressed {
                background-color: #ddd6fe;
            }
            QPushButton:disabled {
                background-color: #f3f4f6;
                border: 1px solid #d1d5db;
                color: #9ca3af;
            }
        """)
        self.btn_import.clicked.connect(self.import_strategy)
        exec_layout.addWidget(self.btn_import, 1, 0)

        # 整表复制：绿色主题
        self.btn_copy = QPushButton("整表复制")
        self.btn_copy.setMinimumHeight(96)
        self.btn_copy.setStyleSheet("""
            QPushButton {
                background-color: #ecfdf5;
                border: 2px solid #10b981;
                border-radius: 10px;
                color: #065f46;
                font-weight: bold;
                padding: 8px 14px;

            }
            QPushButton:hover {
                background-color: #d1fae5;
                border: 2px solid #059669;
            }
            QPushButton:pressed {
                background-color: #a7f3d0;
                border: 2px solid #047857;
            }
            QPushButton:disabled {
                background-color: #f3f4f6;
                border: 2px solid #d1d5db;
                color: #9ca3af;
            }
        """)
        self.btn_copy.clicked.connect(self.do_copy)
        exec_layout.addWidget(self.btn_copy, 0, 1, 2, 1)

        # 设置列拉伸
        exec_layout.setColumnStretch(0, 1)
        exec_layout.setColumnStretch(1, 1)
        exec_layout.setRowStretch(0, 1)
        exec_layout.setRowStretch(1, 1)

        top_row.addWidget(exec_panel, 3)  # 操作区占 30%

        root.addLayout(top_row)

        # ---- 映射配置区 ----
        map_group = QGroupBox("映射配置（每行一条规则）")
        map_layout = QVBoxLayout(map_group)

        self.map_table = MappingTable()
        self.map_table.setColumnCount(MAP_COL_COUNT)
        self.map_table.setHorizontalHeaderLabels(
            ["源表", "填入方式", "源列", "目标列", "连接符", "拆分数", "操作"]
        )
        hdr = self.map_table.horizontalHeader()
        hdr.setSectionResizeMode(QHeaderView.Stretch)
        self.map_table.setSelectionBehavior(QAbstractItemView.SelectRows)
        self.map_table.rowReorderRequested.connect(self._reorder_rows)
        map_layout.addWidget(self.map_table)

        root.addWidget(map_group, 1)

        tabs.addTab(mapping_widget, _tab_icon("#2f6db5"), "表格映射")

        # ---- 表格合并 标签页 ----
        merge_widget = self._build_merge_tab()
        tabs.addTab(merge_widget, _tab_icon("#2e8b57"), "表格合并")

        # 仅给「标签页」上色（页面下方保持默认白底，不做整页染色）
        tabs.setStyleSheet(
            """
            QTabWidget::pane { border: 1px solid #c5c5c5; top: -1px; }
            QTabBar::tab {
                padding: 10px 26px;
                border: 1px solid #c5c5c5; border-bottom: none;
                border-top-left-radius: 6px; border-top-right-radius: 6px;
            }
            QTabBar::tab:selected {
                background: #3a7bd5; color: #ffffff; font-weight: bold;
            }
            QTabBar::tab:!selected { background: #e8e8e8; color: #555555; }
            """
        )
        # 切换标签时刷新底部状态栏：合并标签使用自己的状态行
        tabs.currentChanged.connect(self._on_tab_changed)
        outer.addWidget(tabs)

        # 统一字体：所有按钮、标签页使用同一字号，确保视觉一致
        _btn_font = QFont()
        _btn_font.setPointSize(10)
        for btn in (self.btn_export, self.btn_import, self.btn_copy,
                    self.btn_merge_run, self.btn_merge_save,
                    self.btn_load_source, self.btn_load_target,
                    self.btn_merge_add, self.btn_merge_remove):
            btn.setFont(_btn_font)
        tabs.tabBar().setFont(_btn_font)

    # ========== 表格合并 标签页构建 ==========
    def _build_merge_tab(self):
        w = QWidget()
        v = QVBoxLayout(w)
        v.setSpacing(10)
        v.setContentsMargins(12, 12, 12, 12)

        # ---- 文件选择 ----
        fg = QGroupBox("文件选择")
        fl = QVBoxLayout(fg)
        frow = QHBoxLayout()
        self.btn_merge_add = QPushButton("添加 WPS 表格")
        self.btn_merge_add.clicked.connect(self._merge_add_files)
        self.btn_merge_remove = QPushButton("移除选中")
        self.btn_merge_remove.clicked.connect(self._merge_remove_selected)
        frow.addWidget(self.btn_merge_add)
        frow.addWidget(self.btn_merge_remove)
        frow.addStretch(1)
        fl.addLayout(frow)
        self.merge_list = QListWidget()
        fl.addWidget(self.merge_list)
        v.addWidget(fg)

        # ---- 合并设置 ----
        sg = QGroupBox("合并设置")
        sg.setStyleSheet("""
            QGroupBox {
                font-weight: bold;
            }
        """)
        sl = QFormLayout(sg)
        sl.setVerticalSpacing(12)
        sl.setHorizontalSpacing(10)
        sl.setContentsMargins(14, 14, 14, 14)
        sl.setLabelAlignment(Qt.AlignRight)
        self.spin_merge_header = QSpinBox()
        self.spin_merge_header.setMinimum(1)
        self.spin_merge_header.setMaximum(50)
        self.spin_merge_header.setValue(1)
        sl.addRow("表头行:", self.spin_merge_header)
        self.combo_merge_col = QComboBox()
        self.combo_merge_col.addItem("请先添加文件")
        self.combo_merge_col.setEnabled(False)
        sl.addRow("排序列:", self.combo_merge_col)
        rbox = QHBoxLayout()
        rbox.setSpacing(16)
        self.radio_asc = QRadioButton("升序")
        self.radio_desc = QRadioButton("降序")
        self.radio_asc.setChecked(True)
        rbox.addWidget(self.radio_asc)
        rbox.addWidget(self.radio_desc)
        rbox.addStretch(1)
        sl.addRow("排序方式:", rbox)
        sg.setMinimumWidth(280)

        sgl = QHBoxLayout()
        sgl.addWidget(sg)
        sgl.addSpacing(20)
        # 「合并并排序」+「保存结果」按钮放在合并设置框的右边（垂直排列，居中）
        vbtn = QVBoxLayout()
        vbtn.addStretch(1)
        vbtn.setSpacing(12)
        vbtn.setContentsMargins(20, 0, 0, 0)
        self.btn_merge_run = QPushButton("合并并排序")
        self.btn_merge_run.setMinimumHeight(55)
        self.btn_merge_run.setStyleSheet("""
            QPushButton {
                background-color: transparent;
                border: 2px solid #66bb6a;
                border-radius: 10px;
                color: #2e7d32;
                font-weight: bold;
                padding: 8px 10px;
            }
            QPushButton:hover {
                background-color: #f1f8e9;
                border: 2px solid #43a047;
            }
            QPushButton:pressed {
                background-color: #e8f5e9;
                border: 2px solid #2e7d32;
            }
            QPushButton:disabled {
                background-color: transparent;
                border: 2px solid #c0c4cc;
                color: #999;
            }
        """)
        self.btn_merge_run.clicked.connect(self._merge_and_sort)
        vbtn.addWidget(self.btn_merge_run)
        self.btn_merge_save = QPushButton("保存结果")
        self.btn_merge_save.setMinimumHeight(55)
        self.btn_merge_save.setEnabled(False)
        self.btn_merge_save.clicked.connect(self._merge_save)
        self.btn_merge_save.setStyleSheet("""
            QPushButton {
                background-color: transparent;
                border: 2px solid #3b82f6;
                border-radius: 10px;
                color: #1e40af;
                font-weight: bold;
                padding: 8px 10px;
            }
            QPushButton:hover {
                background-color: #eff6ff;
                border: 2px solid #2563eb;
            }
            QPushButton:pressed {
                background-color: #dbeafe;
                border: 2px solid #1d4ed8;
            }
            QPushButton:disabled {
                background-color: transparent;
                border: 2px solid #c0c4cc;
                color: #999;
            }
        """)
        vbtn.addWidget(self.btn_merge_save)
        vbtn.addStretch(1)
        sgl.addLayout(vbtn)
        sgl.addStretch(1)
        v.addLayout(sgl)

        # ---- 合并预览（保留表头与表头上方内容）----
        pg = QGroupBox("合并预览")
        pl = QVBoxLayout(pg)
        self.merge_preview = QTableWidget()
        self.merge_preview.setEditTriggers(QAbstractItemView.NoEditTriggers)
        self.merge_preview.setMinimumHeight(400)
        self.merge_preview.horizontalHeader().setVisible(False)
        pl.addWidget(self.merge_preview)
        v.addWidget(pg, 1)

        # ---- 状态栏 ----
        self.lbl_merge_status = QLabel("操作状态：等待操作")
        v.addWidget(self.lbl_merge_status)

        # 表头行变化后，若已添加文件则按新行重新读取
        self.spin_merge_header.valueChanged.connect(
            lambda: self._merge_reload() if self.merge_files else None
        )
        return w

    # ========== 表格合并 逻辑 ==========
    def _read_sheet_any(self, path, header_row):
        """读取任意支持的 Excel 文件，返回 (表头列表, 数据行列表, 表头上方行列表)。"""
        ext = path.lower()
        if ext.endswith((".xlsx", ".xlsm")):
            wb = openpyxl.load_workbook(path, data_only=True)
            ws = wb.active
            max_row = ws.max_row
            if header_row < 1 or header_row > max_row:
                raise ValueError(
                    f"表头行 {header_row} 超出范围（共 {max_row} 行）：{os.path.basename(path)}"
                )
            raw = self._merged_row_values(ws, header_row, ws.max_column)
            headers = [self._clean_header(c, i) for i, c in enumerate(raw)]
            data = [list(r) for r in ws.iter_rows(min_row=header_row + 1, values_only=True)]
            above = (
                [list(r) for r in ws.iter_rows(min_row=1, max_row=header_row - 1, values_only=True)]
                if header_row > 1 else []
            )
            return headers, data, above
        elif ext.endswith(".xls"):
            return self.read_xls(path, header_row)
        else:
            raise ValueError(f"不支持的格式：{os.path.basename(path)}")

    def _merge_add_files(self):
        paths, _ = QFileDialog.getOpenFileNames(
            self, "选择 Excel 文件", "",
            "Excel 文件 (*.xlsx *.xls *.xlsm)"
        )
        if not paths:
            return
        for p in paths:
            if p not in self.merge_files:
                self.merge_files.append(p)
        self._refresh_merge_list()
        self._merge_reload()

    def _merge_remove_selected(self):
        for it in self.merge_list.selectedItems():
            p = it.text()
            if p in self.merge_files:
                self.merge_files.remove(p)
        self._refresh_merge_list()
        self._merge_reload()

    def _refresh_merge_list(self):
        self.merge_list.clear()
        for p in self.merge_files:
            self.merge_list.addItem(p)

    def _merge_reload(self):
        """按当前表头行重新读取所有已添加文件，并刷新排序列下拉框。"""
        hr = self.spin_merge_header.value()
        self.merge_sheets = []
        failed = []
        for p in self.merge_files:
            try:
                headers, data, above = self._read_sheet_any(p, hr)
            except Exception as e:
                failed.append(f"{os.path.basename(p)}：{e}")
                continue
            self.merge_sheets.append(
                {"path": p, "headers": headers, "data": data, "above": above}
            )
        if failed:
            QMessageBox.warning(self, "读取失败", "以下文件读取异常，已跳过：\n" + "\n".join(failed))
        if self.merge_sheets:
            self.merge_canonical_headers = self.merge_sheets[0]["headers"]
            # 表头上方内容（如表名）取首个文件，用于预览/保存时保留
            self.merge_title_rows = self.merge_sheets[0].get("above", [])
            self.combo_merge_col.setEnabled(True)
            self.combo_merge_col.clear()
            self.combo_merge_col.addItems(self.merge_canonical_headers)
        else:
            self.merge_canonical_headers = []
            self.merge_title_rows = []
            self.combo_merge_col.setEnabled(False)
            self.combo_merge_col.clear()
            self.combo_merge_col.addItem("（请先添加文件）")
        self.merge_result = None
        self.btn_merge_save.setEnabled(False)
        n = len(self.merge_sheets)
        self.lbl_merge_status.setText(f"操作状态：已加载 {n} 个文件，共 {sum(len(s['data']) for s in self.merge_sheets)} 行数据")

    def _merge_and_sort(self):
        if not self.merge_sheets:
            QMessageBox.warning(self, "提示", "请先添加至少一个 Excel 文件。")
            return
        col = self.combo_merge_col.currentText()
        if not col or col == "（请先添加文件）":
            QMessageBox.warning(self, "提示", "请选择用于合并/排序的列。")
            return
        canon = self.merge_canonical_headers
        if col not in canon:
            QMessageBox.warning(self, "列名不存在", f"所选列「{col}」不在表头中。")
            return
        ci = canon.index(col)

        combined = []
        empty_files = []
        for sh in self.merge_sheets:
            h = sh["headers"]
            d = sh["data"]
            if not d:
                empty_files.append(os.path.basename(sh["path"]))
                continue
            # 按规范表头对齐（用列名映射；缺失列补 None，多余列丢弃）
            idx_map = {name: h.index(name) for name in canon if name in h}
            for row in d:
                newrow = []
                for name in canon:
                    j = idx_map.get(name)
                    newrow.append(row[j] if (j is not None and j < len(row)) else None)
                combined.append(newrow)
        if empty_files:
            QMessageBox.information(
                self, "提示", "以下文件无数据行，已跳过：\n" + "\n".join(empty_files)
            )
        if not combined:
            QMessageBox.warning(self, "无数据", "所有文件都没有可合并的数据行。")
            self.lbl_merge_status.setText("操作状态：无可合并数据")
            return

        asc = self.radio_asc.isChecked()
        try:
            combined.sort(key=lambda r: _sort_key(r[ci]), reverse=not asc)
        except Exception as e:
            QMessageBox.warning(self, "排序失败", f"按列「{col}」排序出错：{e}")
            return

        self.merge_result = (canon, combined)
        self._show_merge_preview(self.merge_title_rows, canon, combined)
        self.btn_merge_save.setEnabled(True)
        self.lbl_merge_status.setText(
            f"合并完成，共 {len(combined)} 行数据（按「{col}」{'升序' if asc else '降序'}）"
        )

    def _show_merge_preview(self, title_rows, headers, rows):
        """在预览表格中展示合并结果：先表头上方内容（如表名），再表头，再数据。

        关键修复：不再用 setHorizontalHeaderLabels（那会把表头固定在表格最顶部，
        导致合并标题行被压到表头下方）。改为隐藏列头，将标题行、表头行、数据行
        全部以普通行的方式按正确顺序写入表格内容区。
        """
        t = self.merge_preview
        # 清除上次的跨列合并记录（clear() 不清理 span，需手动还原）
        for (sr, sc) in self._merge_preview_spans:
            t.setSpan(sr, sc, 1, 1)
        self._merge_preview_spans = []
        t.clear()
        ncols = max(1, len(headers))
        total = len(title_rows) + 1 + len(rows)
        t.setColumnCount(ncols)
        t.setRowCount(total)
        t.horizontalHeader().setVisible(False)

        r = 0
        # ---- 表头上方内容（合并标题 / 表名）----
        for tr in title_rows:
            vals = [tr[c] if c < len(tr) else None for c in range(ncols)]
            non_empty = [v for v in vals
                         if v is not None and str(v).strip() != ""]
            if len(non_empty) == 1 and ncols > 1:
                # 只有一个非空值 → 视为合并标题行，跨全部列合并居中显示
                t.setSpan(r, 0, 1, ncols)
                self._merge_preview_spans.append((r, 0))
                item = QTableWidgetItem(str(non_empty[0]))
                item.setForeground(QColor("#888888"))
                item.setTextAlignment(Qt.AlignCenter)
                f = item.font()
                f.setBold(True)
                item.setFont(f)
                t.setItem(r, 0, item)
            else:
                for c in range(ncols):
                    val = vals[c]
                    item = QTableWidgetItem("" if val is None else str(val))
                    item.setForeground(QColor("#888888"))
                    item.setTextAlignment(Qt.AlignCenter)
                    t.setItem(r, c, item)
            r += 1

        # ---- 表头行（紧跟标题行之后，数据行之前）----
        for c in range(ncols):
            val = headers[c] if c < len(headers) else ""
            item = QTableWidgetItem(str(val))
            f = item.font()
            f.setBold(True)
            item.setFont(f)
            item.setBackground(QColor("#e8eef5"))
            item.setForeground(QColor("#333333"))
            item.setTextAlignment(Qt.AlignCenter)
            t.setItem(r, c, item)
        r += 1

        # ---- 数据行 ----
        for row in rows:
            for c in range(ncols):
                val = row[c] if c < len(row) else None
                t.setItem(r, c, QTableWidgetItem("" if val is None else str(val)))
            r += 1
        t.resizeColumnsToContents()

    def _merge_save(self):
        if not self.merge_result:
            QMessageBox.warning(self, "提示", "请先执行「合并并排序」。")
            return
        headers, rows = self.merge_result
        ts = datetime.datetime.now().strftime("%Y%m%d_%H%M")
        default_name = f"合并表格result{ts}.xlsx"
        out, _ = QFileDialog.getSaveFileName(
            self, "保存合并结果", default_name, "Excel 文件 (*.xlsx)"
        )
        if not out:
            return
        if not out.lower().endswith(".xlsx"):
            out += ".xlsx"
        try:
            wb = openpyxl.Workbook()
            ws = wb.active
            # 保留表头上方内容（如表名）与表头
            for tr in self.merge_title_rows:
                ws.append(list(tr))
            ws.append(headers)
            for r in rows:
                ws.append(r)
            wb.save(out)
            QMessageBox.information(self, "完成", f"已保存至：\n{out}")
            self.lbl_merge_status.setText(f"已保存：{os.path.basename(out)}")
        except Exception as e:
            QMessageBox.critical(self, "保存失败", f"无法保存文件：{e}")

    # ========== 多源表格管理 ==========
    def _add_source_row(self, primary=False):
        """添加一个源表行（UI + 状态）。primary=True 时为主源表（不可删除）。"""
        if len(self.sources) >= self.MAX_SOURCES:
            QMessageBox.information(self, "提示", f"最多支持 {self.MAX_SOURCES} 个源表格。")
            return
        idx = len(self.sources)

        # ---- UI 行 ----
        row_widget = QWidget()
        row_layout = QHBoxLayout(row_widget)
        row_layout.setContentsMargins(4, 4, 4, 4)
        row_layout.setSpacing(8)

        # 上传按钮（带浅蓝底色），宽度与"上传目标表格"一致
        btn_load = QPushButton(f"上传源表格{idx + 1}")
        btn_load.setMinimumWidth(120)
        btn_load.setMinimumHeight(36)
        btn_load.setStyleSheet("""
            QPushButton {
                background-color: #eff6ff;
                border: 1px solid #2563eb;
                border-radius: 6px;
                color: #1e40af;
                font-weight: bold;
                padding: 4px 12px;
                text-align: center;
            }
            QPushButton:hover {
                background-color: #dbeafe;
                border: 1px solid #1d4ed8;
            }
            QPushButton:pressed {
                background-color: #bfdbfe;
            }
        """)
        row_layout.addWidget(btn_load)

        # 文件名（固定宽度，确保所有行的子表下拉框左对齐）
        lbl_name = QLabel("（未选择）")
        lbl_name.setWordWrap(False)
        lbl_name.setFixedWidth(450)
        row_layout.addWidget(lbl_name)

        # 子表选择（10 个汉字宽度）
        lbl_sheet = QLabel("子表:")
        row_layout.addWidget(lbl_sheet)
        combo_sheet = QComboBox()
        combo_sheet.setEnabled(False)
        combo_sheet.setMinimumWidth(200)   # 10 个汉字宽度
        combo_sheet.setToolTip("选择子表（Sheet）")
        row_layout.addWidget(combo_sheet)
        row_layout.addSpacing(12)   # 子表后预留空间

        # 表头行（显示"第 N 行"格式，5 个汉字宽度）
        lbl_hr = QLabel("表头行:")
        row_layout.addWidget(lbl_hr)
        spin_header = QSpinBox()
        spin_header.setMinimum(1)
        spin_header.setMaximum(50)
        spin_header.setValue(1)
        spin_header.setAlignment(Qt.AlignCenter)
        spin_header.setMinimumWidth(90)   # 5 个汉字宽度
        spin_header.setPrefix("第 ")
        spin_header.setSuffix(" 行")
        row_layout.addWidget(spin_header)
        row_layout.addSpacing(12)   # 表头行后预留空间

        # 一键匹配按钮
        btn_match = QPushButton("一键匹配表头")
        btn_match.setMinimumWidth(120)
        btn_match.setMinimumHeight(36)
        btn_match.setCursor(Qt.PointingHandCursor)
        btn_match.setStyleSheet("""
            QPushButton {
                background-color: #ecfdf5;
                border: 1px solid #10b981;
                border-radius: 6px;
                color: #065f46;
                font-weight: bold;
                padding: 4px 12px;
                text-align: center;
            }
            QPushButton:hover {
                background-color: #d1fae5;
                border: 1px solid #059669;
            }
            QPushButton:pressed {
                background-color: #a7f3d0;
            }
            QPushButton:disabled {
                background-color: #f3f4f6;
                border: 1px solid #d1d5db;
                color: #9ca3af;
            }
        """)
        row_layout.addWidget(btn_match)
        row_layout.addSpacing(12)   # 一键匹配与删除之间预留空间

        if not primary:
            btn_del = QPushButton("删除")
            btn_del.setMinimumWidth(96)   # 4 个汉字宽度
            btn_del.setMinimumHeight(36)
            btn_del.setStyleSheet("""
                QPushButton {
                    background-color: #fef2f2;
                    border: 1px solid #ef4444;
                    border-radius: 6px;
                    color: #991b1b;
                    font-weight: bold;
                    padding: 4px 10px;
                    text-align: center;
                }
                QPushButton:hover {
                    background-color: #fee2e2;
                    border: 1px solid #dc2626;
                }
                QPushButton:pressed {
                    background-color: #fecaca;
                }
            """)
            row_layout.addWidget(btn_del)
        else:
            btn_del = None

        row_layout.addStretch(1)

        self.src_container.addWidget(row_widget)

        # ---- 状态 ----
        src_info = {
            "path": "",
            "sheet": "",
            "headers": [],
            "data": [],
            "header_row": 1,
        }
        self.sources.append(src_info)
        w_info = {
            "row_widget": row_widget,
            "btn_load": btn_load,
            "lbl_name": lbl_name,
            "combo_sheet": combo_sheet,
            "spin_header": spin_header,
            "btn_match": btn_match,
            "btn_del": btn_del,
            "primary": primary,
        }
        self.source_widgets.append(w_info)

        # 信号绑定
        src_idx = idx  # 闭包捕获
        btn_load.clicked.connect(lambda _=None, i=src_idx: self._load_source_at(i))
        if btn_del:
            btn_del.clicked.connect(lambda _=None, i=src_idx: self._remove_source_at(i))
        combo_sheet.currentIndexChanged.connect(
            lambda _, i=src_idx: self._on_source_sheet_changed_at(i)
        )
        spin_header.valueChanged.connect(
            lambda _=None, i=src_idx: self._on_source_header_row_changed_at(i)
        )
        btn_match.clicked.connect(lambda _=None, i=src_idx: self._auto_match_headers_for(i))

        # 主源表：绑定到旧变量名（向后兼容）
        if primary:
            self.btn_load_source = btn_load
            self.lbl_source = lbl_name
            self.combo_source_sheet = combo_sheet
            self.spin_source_header = spin_header

        # 更新"添加源表格"按钮状态
        self.btn_add_source.setEnabled(len(self.sources) < self.MAX_SOURCES)

        # 更新所有映射行的源表下拉框
        self._refresh_srctbl_combos()

    def _remove_source_at(self, idx, silent=False):
        """删除非主源表行。silent=True 时跳过确认弹窗。"""
        if idx == 0:
            return  # 主源表不可删除
        if idx >= len(self.sources):
            return
        if not silent:
            # 确认
            src = self.sources[idx]
            fname = os.path.basename(src.get("path", "")) if src.get("path") else f"源表{idx + 1}"
            reply = QMessageBox.question(
                self, "确认删除",
                f"确定要删除「{fname}」吗？\n相关映射行中的源表选择会被重置为源表1。",
                QMessageBox.Yes | QMessageBox.No, QMessageBox.No,
            )
            if reply != QMessageBox.Yes:
                return

        # 移除 UI
        w = self.source_widgets[idx]
        w["row_widget"].deleteLater()
        self.source_widgets.pop(idx)
        self.sources.pop(idx)

        # 更新后续源表的按钮文字和信号
        for i in range(idx, len(self.source_widgets)):
            wi = self.source_widgets[i]
            # 更新上传按钮文字
            wi["btn_load"].setText(f"上传源表格{i + 1}")
            # 重新绑定信号（闭包索引已固定，需要断开重连）
            wi["btn_load"].clicked.disconnect()
            wi["combo_sheet"].currentIndexChanged.disconnect()
            wi["spin_header"].valueChanged.disconnect()
            wi["btn_match"].clicked.disconnect()
            if wi.get("btn_del"):
                wi["btn_del"].clicked.disconnect()
            new_idx = i
            wi["btn_load"].clicked.connect(lambda _=None, ii=new_idx: self._load_source_at(ii))
            wi["combo_sheet"].currentIndexChanged.connect(
                lambda _, ii=new_idx: self._on_source_sheet_changed_at(ii)
            )
            wi["spin_header"].valueChanged.connect(
                lambda _=None, ii=new_idx: self._on_source_header_row_changed_at(ii)
            )
            wi["btn_match"].clicked.connect(lambda _=None, ii=new_idx: self._auto_match_headers_for(ii))
            if wi.get("btn_del"):
                wi["btn_del"].clicked.connect(lambda _=None, ii=new_idx: self._remove_source_at(ii))

        # 更新映射行中源表下拉框
        self._refresh_srctbl_combos()

        # 重置映射行中已删除源表的选中项为源表1
        for r in range(self.map_table.rowCount()):
            cb = self.map_table.cellWidget(r, COL_SRCTBL)
            if cb and cb.currentIndex() >= len(self.sources):
                cb.setCurrentIndex(0)

        self.btn_add_source.setEnabled(len(self.sources) < self.MAX_SOURCES)
        self.statusBar().showMessage(f"已删除源表，当前共 {len(self.sources)} 个源表")

    def _source_label(self, idx):
        """生成源表下拉框的显示文本。"""
        src = self.sources[idx] if idx < len(self.sources) else None
        if not src or not src.get("path"):
            return f"源表{idx + 1}（未加载）"
        fname = os.path.basename(src["path"])
        return f"源表{idx + 1}: {fname}"

    def _refresh_srctbl_combos(self):
        """刷新所有映射行的源表下拉框选项。"""
        if not hasattr(self, "map_table"):
            return  # 映射表尚未创建（初始化阶段）
        labels = [self._source_label(i) for i in range(len(self.sources))]
        for r in range(self.map_table.rowCount()):
            cb = self.map_table.cellWidget(r, COL_SRCTBL)
            if not cb:
                continue
            cur = cb.currentText()
            cb.blockSignals(True)
            cb.clear()
            cb.addItems(labels)
            # 尝试保持原选择
            if cur in labels:
                cb.setCurrentText(cur)
            else:
                cb.setCurrentIndex(0)
            cb.blockSignals(False)
            # 源表切换后重建源列控件
            self._on_srctbl_changed_at(r)

    def _on_srctbl_changed_at(self, row):
        """映射行中源表下拉框变化时，重建该行的源列控件。"""
        cb = self.map_table.cellWidget(row, COL_SRCTBL)
        if not cb:
            return
        src_idx = cb.currentIndex()
        if src_idx < 0 or src_idx >= len(self.sources):
            return
        method_w = self.map_table.cellWidget(row, COL_METHOD)
        method = method_w.currentText() if method_w else "直接复制"
        self.rebuild_source_cell(row, method, src_idx)

    def _load_source_at(self, idx):
        """为指定源表打开文件对话框并加载。"""
        if idx >= len(self.sources):
            return
        path, _ = QFileDialog.getOpenFileName(
            self, f"选择源表{idx + 1}", "",
            "Excel 文件 (*.xlsx *.xls *.xlsm)"
        )
        if not path:
            return
        src = self.sources[idx]
        src["path"] = path
        src["header_row"] = self.source_widgets[idx]["spin_header"].value()
        w = self.source_widgets[idx]
        w["lbl_name"].setText(os.path.basename(path))
        w["lbl_name"].setToolTip(path)
        # 填充子表
        self._populate_source_sheets_at(idx)
        # 加载数据
        self._apply_source_load_at(idx)
        # 刷新映射行源表下拉框（显示文件名）
        self._refresh_srctbl_combos()

    def _populate_source_sheets_at(self, idx):
        """填充指定源表的子表下拉框。"""
        if idx >= len(self.sources):
            return
        src = self.sources[idx]
        path = src.get("path", "")
        w = self.source_widgets[idx]
        combo = w["combo_sheet"]
        if not path:
            combo.clear()
            combo.setEnabled(False)
            src["sheet"] = ""
            return
        ext = path.lower()
        try:
            sheet_names = []
            if ext.endswith((".xlsx", ".xlsm")):
                wb = openpyxl.load_workbook(path, read_only=True)
                sheet_names = wb.sheetnames
                wb.close()
            elif ext.endswith(".xls"):
                try:
                    import xlrd
                    wb = xlrd.open_workbook(path)
                    sheet_names = wb.sheet_names()
                except ImportError:
                    pass
            if not sheet_names:
                sheet_names = ["Sheet1"]
        except Exception:
            sheet_names = ["Sheet1"]

        combo.blockSignals(True)
        combo.clear()
        combo.addItems(sheet_names)
        combo.setEnabled(len(sheet_names) > 1)
        src["sheet"] = sheet_names[0]
        combo.setCurrentIndex(0)
        combo.blockSignals(False)

    def _on_source_sheet_changed_at(self, idx):
        """指定源表的子表切换时重新加载数据。"""
        if idx >= len(self.sources):
            return
        w = self.source_widgets[idx]
        combo = w["combo_sheet"]
        sheet_name = combo.currentText()
        src = self.sources[idx]
        if not src.get("path") or sheet_name == src.get("sheet"):
            return
        src["sheet"] = sheet_name
        self._apply_source_load_at(idx)
        # 刷新引用此源表的映射行
        self._refresh_srctbl_combos()

    def _apply_source_load_at(self, idx):
        """按当前表头行设置读取指定源表的表头与数据。"""
        if idx >= len(self.sources):
            return
        src = self.sources[idx]
        path = src.get("path", "")
        if not path:
            return
        hr = self.source_widgets[idx]["spin_header"].value()
        ext = path.lower()
        try:
            if ext.endswith((".xlsx", ".xlsm")):
                wb = openpyxl.load_workbook(path, data_only=True)
                ws = wb[src["sheet"]] if src.get("sheet") else wb.active
                max_row = ws.max_row
                max_col = ws.max_column
                if hr < 1 or hr > max_row:
                    QMessageBox.warning(
                        self, "表头行错误",
                        f"源表{idx + 1} 表头行 {hr} 超出范围（共 {max_row} 行）。"
                    )
                    wb.close()
                    return
                raw = self._merged_row_values(ws, hr, max_col)
                headers = [self._clean_header(c, i) for i, c in enumerate(raw)]
                data = [
                    list(r)
                    for r in ws.iter_rows(min_row=hr + 1, values_only=True)
                ]
                wb.close()
            elif ext.endswith(".xls"):
                headers, data, _ = self.read_xls(path, hr, src.get("sheet"))
            else:
                QMessageBox.warning(self, "格式错误", f"源表{idx + 1}：不支持的文件格式。")
                return
        except ImportError as e:
            QMessageBox.warning(self, "缺少依赖", str(e))
            return
        except Exception as e:
            QMessageBox.critical(self, "读取失败", f"无法读取源表{idx + 1}：{e}")
            return

        src["headers"] = headers
        src["data"] = data
        src["header_row"] = hr

        # 主源表（idx=0）：同步到旧变量（向后兼容）
        if idx == 0:
            self.source_path = path
            self.source_headers = headers
            self.source_data = data
            self.source_header_row = hr
            self.source_sheet_name = src.get("sheet", "")

        self.statusBar().showMessage(
            f"已加载源表{idx + 1}：{len(headers)}列，{len(data)}行数据（表头第 {hr} 行）"
        )
        self.refresh_after_load()

    def _get_source_idx_for_row(self, row):
        """获取映射行当前选中的源表索引。"""
        cb = self.map_table.cellWidget(row, COL_SRCTBL)
        if not cb:
            return 0
        idx = cb.currentIndex()
        if idx < 0 or idx >= len(self.sources):
            return 0
        return idx

    def _get_headers_for_row(self, row):
        """获取映射行当前源表的表头列表。"""
        idx = self._get_source_idx_for_row(row)
        return self.sources[idx].get("headers", []) if idx < len(self.sources) else []

    def _on_source_header_row_changed_at(self, idx):
        """指定源表的表头行变化时，重新加载该源表。"""
        if idx < len(self.sources) and self.sources[idx].get("path"):
            self._apply_source_load_at(idx)

    # ========== 映射行管理 ==========
    def _init_mapping(self, rows):
        self.map_table.setRowCount(rows)
        for r in range(rows):
            self.add_mapping_row(r)

    def add_mapping_row(self, row, default_method="直接复制", config=None):
        # 导入配置时，用 config 中的 method 决定控件类型（否则合并/拆分的控件会建错）
        if config and config.get("method"):
            default_method = config["method"]

        # 源表选择（第一列）
        srctbl_cb = QComboBox()
        srctbl_cb.setStyleSheet("QComboBox { text-align: center; } QComboBox::item { text-align: center; }")
        srctbl_cb.currentIndexChanged.connect(lambda _: self._on_srctbl_changed_at(row))
        self.map_table.setCellWidget(row, COL_SRCTBL, srctbl_cb)

        # 填入方式
        mc = QComboBox()
        mc.addItems(["直接复制", "合并复制", "拆分复制"])
        mc.setCurrentText(default_method)
        mc.currentTextChanged.connect(lambda _=None: self.on_method_changed())
        mc.setStyleSheet("QComboBox { text-align: center; } QComboBox::item { text-align: center; }")
        mc.setEditable(True)
        mc.lineEdit().setAlignment(Qt.AlignCenter)
        mc.lineEdit().setReadOnly(True)
        self.map_table.setCellWidget(row, COL_METHOD, mc)

        # 源列（依据方式决定控件类型）
        self.rebuild_source_cell(row, default_method, 0)

        # 目标列（依据方式决定控件类型）
        self.rebuild_target_cell(row, default_method)

        # 连接符
        sep = QLineEdit("-")
        sep.setEnabled(default_method == "合并复制")
        sep.setAlignment(Qt.AlignCenter)
        self.map_table.setCellWidget(row, COL_SEP, sep)

        # 拆分数
        sp = QSpinBox()
        sp.setMinimum(1)
        sp.setMaximum(max(1, self.target_col_count))
        sp.setValue(2 if self.target_col_count >= 2 else 1)
        sp.setEnabled(default_method == "拆分复制")
        sp.setAlignment(Qt.AlignCenter)
        self.map_table.setCellWidget(row, COL_SPLIT, sp)

        # 操作
        aw = QWidget()
        hl = QHBoxLayout(aw)
        hl.setContentsMargins(2, 2, 2, 2)
        bplus = QPushButton("增行")
        bminus = QPushButton("删行")
        bplus.setFixedWidth(72)
        bminus.setFixedWidth(72)
        bplus.setStyleSheet("QPushButton { text-align: center; }")
        bminus.setStyleSheet("QPushButton { text-align: center; }")
        bplus.clicked.connect(lambda _=None: self.on_plus())
        bminus.clicked.connect(lambda _=None: self.on_minus())
        hl.addWidget(bplus)
        hl.addWidget(bminus)
        self.map_table.setCellWidget(row, COL_ACTION, aw)

        # 填充源表下拉框选项
        labels = [self._source_label(i) for i in range(len(self.sources))]
        srctbl_cb.blockSignals(True)
        srctbl_cb.clear()
        srctbl_cb.addItems(labels)
        if config and config.get("src_tbl_idx") is not None:
            idx = config["src_tbl_idx"]
            if 0 <= idx < len(labels):
                srctbl_cb.setCurrentIndex(idx)
        else:
            srctbl_cb.setCurrentIndex(0)
        srctbl_cb.blockSignals(False)
        # 源表选定后重建源列控件
        self._on_srctbl_changed_at(row)

        # 应用拖拽/重建时传入的配置（不传则保持空白默认）
        if config:
            self._apply_source(row, config.get("src"))
            self._apply_target(row, config.get("tgt"))
            sep.setText(config.get("sep") or "-")
            sp.setValue(config.get("split") or 1)

    def _apply_source(self, row, sel):
        w = self.map_table.cellWidget(row, COL_SOURCE)
        if isinstance(w, QPushButton):   # 合并复制（多选按钮）
            w.setProperty("selected", sel or [])
            w.setText(", ".join(sel) if sel else "（未选择）")
        elif isinstance(w, QComboBox):
            idx = w.findText(sel) if sel else -1
            w.setCurrentIndex(idx if idx >= 0 else -1)

    def _apply_target(self, row, sel):
        w = self.map_table.cellWidget(row, COL_TARGET)
        if isinstance(w, QPushButton):   # 拆分复制（多选按钮）
            w.setProperty("selected", sel or [])
            w.setText(", ".join(sel) if sel else "（未选择）")
        elif isinstance(w, QComboBox):
            idx = w.findText(sel) if sel else -1
            w.setCurrentIndex(idx if idx >= 0 else -1)

    def read_row_config(self, row):
        """读取某行当前配置，供拖拽重排序后重建使用。"""
        method = self.map_table.cellWidget(row, COL_METHOD).currentText()
        sep_w = self.map_table.cellWidget(row, COL_SEP)
        sp_w = self.map_table.cellWidget(row, COL_SPLIT)
        srctbl_w = self.map_table.cellWidget(row, COL_SRCTBL)
        return {
            "method": method,
            "src_tbl_idx": srctbl_w.currentIndex() if srctbl_w else 0,
            "src": self.get_source_selection(row),
            "tgt": self.get_target_selection(row),
            "sep": sep_w.text() if sep_w else "-",
            "split": sp_w.value() if sp_w else 1,
        }

    def _reorder_rows(self, order):
        """按拖拽后的行顺序重建所有行，保留每行全部配置。"""
        n = self.map_table.rowCount()
        if len(order) != n or n <= 1:
            return
        configs = [self.read_row_config(r) for r in range(n)]
        new_configs = [configs[r] for r in order]
        self.map_table.setRowCount(0)   # 删除旧行（含其控件）
        for i, cfg in enumerate(new_configs):
            self.map_table.insertRow(i)
            self.add_mapping_row(i, config=cfg)
        # 重建后表头视觉顺序重置为 0..n-1；确保行号仍可继续拖拽
        self.map_table.verticalHeader().setSectionsMovable(True)
        self.statusBar().showMessage("已按拖拽顺序调整映射行")

    # ========== 配置导出 / 导入 ==========
    def export_strategy(self):
        """把当前所有映射行及表头行设置导出为 JSON 文件。"""
        n = self.map_table.rowCount()
        if n == 0:
            QMessageBox.information(self, "提示", "当前没有可导出的映射行。")
            return
        default_dir = os.path.dirname(self.target_path or self.source_path or ".")
        path, _ = QFileDialog.getSaveFileName(
            self, "导出配置",
            os.path.join(default_dir, "mapping_config.json"),
            "JSON 文件 (*.json)"
        )
        if not path:
            return
        if not path.lower().endswith(".json"):
            path += ".json"
        data = {
            "version": 4,
            "sources": [
                {
                    "path": s.get("path", ""),
                    "sheet": s.get("sheet", ""),
                    "header_row": self.source_widgets[i]["spin_header"].value()
                        if i < len(self.source_widgets) else 1,
                }
                for i, s in enumerate(self.sources)
            ],
            "source_path": self.source_path,       # 向后兼容
            "target_path": self.target_path,
            "source_sheet": self.source_sheet_name,  # 向后兼容
            "target_sheet": self.target_sheet_name,
            "source_header_row": self.spin_source_header.value(),
            "target_header_row": self.spin_target_header.value(),
            "rules": [self.read_row_config(r) for r in range(n)],
        }
        try:
            with open(path, "w", encoding="utf-8") as f:
                json.dump(data, f, ensure_ascii=False, indent=2)
            QMessageBox.information(
                self, "完成",
                f"已导出 {len(data['rules'])} 条映射配置至：\n{path}"
            )
            self.statusBar().showMessage(f"已导出配置：{os.path.basename(path)}")
        except Exception as e:
            QMessageBox.critical(self, "错误", f"导出失败：{e}")

    def import_strategy(self):
        """从 JSON 文件导入配置并重建所有行。"""
        default_dir = os.path.dirname(self.target_path or self.source_path or ".")
        path, _ = QFileDialog.getOpenFileName(
            self, "导入配置", default_dir, "JSON 文件 (*.json)"
        )
        if not path:
            return
        try:
            with open(path, "r", encoding="utf-8") as f:
                data = json.load(f)
        except Exception as e:
            QMessageBox.critical(self, "错误", f"无法读取配置文件：{e}")
            return
        rules = data.get("rules") if isinstance(data, dict) else None
        if not isinstance(rules, list) or not rules:
            QMessageBox.warning(self, "提示", "配置文件为空或格式不正确。")
            return

        # 恢复表头行设置（v3 兼容：source_header_row 适用于主源表）
        if isinstance(data, dict):
            if data.get("source_header_row"):
                self.spin_source_header.setValue(data["source_header_row"])
            if data.get("target_header_row"):
                self.spin_target_header.setValue(data["target_header_row"])

        # 清除现有额外源表（保留主源表行）
        while len(self.sources) > 1:
            self._remove_source_at(len(self.sources) - 1, silent=True)
        # 清除主源表的状态
        self.sources[0] = {"path": "", "sheet": "", "headers": [], "data": [], "header_row": 1}
        self.source_widgets[0]["lbl_name"].setText("（未选择）")
        self.source_widgets[0]["combo_sheet"].clear()
        self.source_widgets[0]["combo_sheet"].setEnabled(False)

        # 恢复源表格（支持多源表 v4 格式，兼容旧版 v3 单源表格式）
        sources_list = data.get("sources", []) if isinstance(data, dict) else []
        if not sources_list and data.get("source_path"):
            # v3 兼容：单源表格式
            sources_list = [{"path": data["source_path"], "sheet": data.get("source_sheet", "")}]

        for i, src_item in enumerate(sources_list[:self.MAX_SOURCES]):
            src_path = src_item.get("path", "")
            src_sheet = src_item.get("sheet", "")
            src_hr = src_item.get("header_row", 1)
            if not src_path or not os.path.isfile(src_path):
                if src_path:
                    QMessageBox.warning(self, "提示", f"源表{i + 1}文件不存在，已跳过：\n{src_path}")
                continue
            if i >= len(self.sources):
                self._add_source_row()
            # 恢复表头行设置（在加载数据前设置）
            self.source_widgets[i]["spin_header"].setValue(src_hr)
            self.sources[i]["path"] = src_path
            self.source_widgets[i]["lbl_name"].setText(os.path.basename(src_path))
            self.source_widgets[i]["lbl_name"].setToolTip(src_path)
            self._populate_source_sheets_at(i)
            if src_sheet:
                combo = self.source_widgets[i]["combo_sheet"]
                idx = combo.findText(src_sheet)
                if idx >= 0:
                    combo.setCurrentIndex(idx)
                    self.sources[i]["sheet"] = src_sheet
            self._apply_source_load_at(i)

        # 恢复目标表格
        tgt_path = data.get("target_path", "") if isinstance(data, dict) else ""
        tgt_sheet = data.get("target_sheet", "") if isinstance(data, dict) else ""
        if tgt_path and os.path.isfile(tgt_path):
            self.target_path = tgt_path
            self.lbl_target.setText(f"目标文件：{os.path.basename(tgt_path)}")
            self._populate_target_sheets()
            if tgt_sheet and self.combo_target_sheet.findText(tgt_sheet) >= 0:
                self.combo_target_sheet.setCurrentText(tgt_sheet)
                self.target_sheet_name = tgt_sheet
            self._apply_target_load()
        elif tgt_path:
            QMessageBox.warning(self, "提示", f"目标表格文件不存在，已跳过：\n{tgt_path}")

        self.map_table.setRowCount(0)
        for i, cfg in enumerate(rules):
            self.map_table.insertRow(i)
            self.add_mapping_row(i, config=cfg)
        self.map_table.verticalHeader().setSectionsMovable(True)
        self._validate_imported_columns(rules)
        QMessageBox.information(
            self, "完成", f"已导入 {len(rules)} 条映射配置。"
        )
        self.statusBar().showMessage(f"已导入配置：{os.path.basename(path)}")

    def _validate_imported_columns(self, rules):
        """导入后若当前表已加载，检查列名是否存在，缺失则提示。"""
        if not self.source_headers or not self.target_headers:
            return
        miss = []
        for r in rules:
            src = r.get("src")
            tgt = r.get("tgt")
            if isinstance(src, list):
                for s in src:
                    if s and s not in self.source_headers:
                        miss.append(("源列", s))
            elif src and src not in self.source_headers:
                miss.append(("源列", src))
            if isinstance(tgt, list):
                for t in tgt:
                    if t and t not in self.target_headers:
                        miss.append(("目标列", t))
            elif tgt and tgt not in self.target_headers:
                miss.append(("目标列", tgt))
        if miss:
            shown = "\n".join(f"[{k}] {v}" for k, v in miss[:20])
            QMessageBox.warning(
                self, "列名不匹配",
                "以下列名在当前表格中找不到（相关行复制时会被跳过）：\n" + shown
            )

    def rebuild_source_cell(self, row, method, src_idx=None):
        old = self.map_table.cellWidget(row, COL_SOURCE)
        if old:
            old.deleteLater()
        # 获取该行当前源表的表头
        if src_idx is None:
            src_idx = self._get_source_idx_for_row(row)
        headers = (self.sources[src_idx].get("headers", [])
                   if src_idx < len(self.sources) else [])
        if method == "合并复制":
            btn = QPushButton("点击选择源列…")
            btn.setProperty("selected", [])
            btn.setProperty("kind", "source")
            btn.setProperty("src_idx", src_idx)
            btn.clicked.connect(lambda _=None: self.open_multi_select())
            self.map_table.setCellWidget(row, COL_SOURCE, btn)
        else:
            cb = QComboBox()
            cb.addItems(headers)
            cb.setCurrentIndex(-1)   # 留空，不预填充
            self.map_table.setCellWidget(row, COL_SOURCE, cb)

    def rebuild_target_cell(self, row, method):
        old = self.map_table.cellWidget(row, COL_TARGET)
        if old:
            old.deleteLater()
        if method == "拆分复制":
            btn = QPushButton("点击选择目标列…")
            btn.setProperty("selected", [])
            btn.setProperty("kind", "target")
            btn.clicked.connect(lambda _=None: self.open_multi_select())
            self.map_table.setCellWidget(row, COL_TARGET, btn)
        else:
            cb = QComboBox()
            cb.addItems(self.target_headers)
            cb.setCurrentIndex(-1)   # 留空，不预填充
            self.map_table.setCellWidget(row, COL_TARGET, cb)

    def on_method_changed(self):
        combo = self.sender()
        row = self._find_row(COL_METHOD, combo)
        if row < 0:
            return
        method = combo.currentText()
        src_idx = self._get_source_idx_for_row(row)
        self.rebuild_source_cell(row, method, src_idx)
        self.rebuild_target_cell(row, method)
        sep = self.map_table.cellWidget(row, COL_SEP)
        sp = self.map_table.cellWidget(row, COL_SPLIT)
        if sep:
            sep.setEnabled(method == "合并复制")
        if sp:
            sp.setEnabled(method == "拆分复制")

    def open_multi_select(self):
        btn = self.sender()
        kind = btn.property("kind")
        col = COL_SOURCE if kind == "source" else COL_TARGET
        row = self._find_row(col, btn)
        if row < 0:
            return
        if kind == "source":
            # 使用该行当前源表的表头
            src_idx = self._get_source_idx_for_row(row)
            items = (self.sources[src_idx].get("headers", [])
                     if src_idx < len(self.sources) else [])
            max_sel = 5
            title = f"选择源列（源表{src_idx + 1}，合并复制，可选 2~5 个）"
        else:
            items = self.target_headers
            sp = self.map_table.cellWidget(row, COL_SPLIT)
            max_sel = sp.value() if isinstance(sp, QSpinBox) else 2
            title = f"选择目标列（拆分复制，需选 {max_sel} 个）"
        dlg = MultiSelectDialog(
            items, btn.property("selected") or [], self,
            max_sel=max_sel, title=title,
        )
        if dlg.exec_() == QDialog.Accepted:
            sel = dlg.selected()
            btn.setProperty("selected", sel)
            btn.setText(", ".join(sel) if sel else "（未选择）")

    def on_plus(self):
        btn = self.sender()
        row = self._find_action_row(btn)
        if row < 0:
            return
        self.map_table.insertRow(row + 1)
        self.add_mapping_row(row + 1)

    def on_minus(self):
        btn = self.sender()
        row = self._find_action_row(btn)
        if row < 0:
            return
        if self.map_table.rowCount() > 1:
            self.map_table.removeRow(row)
        else:
            QMessageBox.information(self, "提示", "至少保留一行映射规则。")

    def _find_row(self, col, widget):
        for r in range(self.map_table.rowCount()):
            if self.map_table.cellWidget(r, col) is widget:
                return r
        return -1

    def _find_action_row(self, btn):
        """+/- 按钮嵌在 COL_ACTION 的容器 QWidget 内，需按子控件定位行。"""
        for r in range(self.map_table.rowCount()):
            aw = self.map_table.cellWidget(r, COL_ACTION)
            if aw is not None and btn in aw.findChildren(QPushButton):
                return r
        return -1

    def get_source_selection(self, row):
        """返回源列选择：直接/拆分为单名(str)，合并为列表(list)。"""
        w = self.map_table.cellWidget(row, COL_SOURCE)
        if isinstance(w, QPushButton):
            return w.property("selected") or []
        return w.currentText() if w else ""

    def get_target_selection(self, row):
        """返回目标列选择：直接/合并为单名(str)，拆分为列表(list)。"""
        w = self.map_table.cellWidget(row, COL_TARGET)
        if isinstance(w, QPushButton):
            return w.property("selected") or []
        return w.currentText() if w else ""

    def refresh_after_load(self):
        # 文件加载后重建所有源/目标列控件（保持空白，不预填充），并更新拆分数上限。
        for row in range(self.map_table.rowCount()):
            method_w = self.map_table.cellWidget(row, COL_METHOD)
            method = method_w.currentText() if method_w else "直接复制"
            src_idx = self._get_source_idx_for_row(row)
            self.rebuild_source_cell(row, method, src_idx)
            self.rebuild_target_cell(row, method)
            sp = self.map_table.cellWidget(row, COL_SPLIT)
            if isinstance(sp, QSpinBox):
                sp.setMaximum(max(1, self.target_col_count))

    # ========== 模糊表头匹配 ==========
    # 标准字段及其匹配规则
    _STANDARD_FIELDS = ["姓名", "性别", "出生年月", "民族"]

    @staticmethod
    def _normalize_header(s):
        """去除空格（全角/半角）、常见标点，统一小写用于匹配。"""
        if not s:
            return ""
        s = str(s).strip().lower()
        # 去掉全角空格、半角空格、常见标点
        for ch in ("\u3000", " ", "\t", ".", "．", "、", "/", "\\", "-", "_",
                   "（", "）", "(", ")", "[", "]", "【", "】", "：", ":"):
            s = s.replace(ch, "")
        return s

    def _fuzzy_find_match(self, standard_field, headers):
        """在给定的表头列表中，为标准字段找到最佳模糊匹配项。

        返回 (best_header, score) 或 (None, 0)。
        score 越大表示匹配度越高。
        """
        norm_headers = [(h, self._normalize_header(h)) for h in headers]
        best = None
        best_score = 0

        if standard_field == "姓名":
            for orig, norm in norm_headers:
                if not norm:
                    continue
                score = 0
                # 精确包含"姓名"
                if "姓名" in norm:
                    score = max(score, 100 - abs(len(norm) - 2) * 5)
                # 同时包含"姓"和"名"（顺序不限）
                if "姓" in norm and "名" in norm:
                    score = max(score, 80 - abs(len(norm) - 2) * 5)
                # 包含"名字"
                if "名字" in norm:
                    score = max(score, 70)
                if score > best_score:
                    best_score = score
                    best = orig

        elif standard_field == "性别":
            for orig, norm in norm_headers:
                if not norm:
                    continue
                score = 0
                # 精确包含"性别"
                if "性别" in norm:
                    score = max(score, 100 - abs(len(norm) - 2) * 5)
                # 同时包含"性"和"别"
                if "性" in norm and "别" in norm:
                    score = max(score, 80 - abs(len(norm) - 2) * 5)
                # 英文/其他标识
                if norm in ("sex", "gender"):
                    score = max(score, 75)
                if "m/f" in norm or "mf" in norm:
                    score = max(score, 60)
                if "男" in norm or "女" in norm:
                    score = max(score, 55)
                if score > best_score:
                    best_score = score
                    best = orig

        elif standard_field == "出生年月":
            for orig, norm in norm_headers:
                if not norm:
                    continue
                score = 0
                # 包含"出生年月"
                if "出生年月" in norm:
                    score = max(score, 100 - abs(len(norm) - 4) * 5)
                # 包含"出生" + "年月"
                if "出生" in norm and ("年月" in norm or "年月日" in norm):
                    score = max(score, 90 - abs(len(norm) - 4) * 3)
                # 包含"出生" + "日期"
                if "出生" in norm and "日期" in norm:
                    score = max(score, 85)
                # 包含"出生" + "时间"
                if "出生" in norm and "时间" in norm:
                    score = max(score, 80)
                # 仅包含"出生"
                if "出生" in norm:
                    score = max(score, 50)
                # 包含"年月"或"日期"或"时间"
                if "年月" in norm or "日期" in norm or "时间" in norm:
                    score = max(score, 40)
                if score > best_score:
                    best_score = score
                    best = orig

        elif standard_field == "民族":
            for orig, norm in norm_headers:
                if not norm:
                    continue
                score = 0
                # 精确包含"民族"
                if "民族" in norm:
                    score = max(score, 100 - abs(len(norm) - 2) * 5)
                # 同时包含"民"和"族"
                if "民" in norm and "族" in norm:
                    score = max(score, 80 - abs(len(norm) - 2) * 5)
                # 包含"族别"
                if "族别" in norm:
                    score = max(score, 55)
                if score > best_score:
                    best_score = score
                    best = orig

        return (best, best_score) if best_score > 0 else (None, 0)

    def _auto_match_headers_for(self, src_idx):
        """为指定源表一键匹配表头：先模糊匹配标准字段，再精确匹配其余相同表头。
        匹配结果以新行追加到映射表（不清除已有配置）。
        """
        if src_idx >= len(self.sources):
            return
        src_headers = self.sources[src_idx].get("headers", [])
        if not src_headers:
            QMessageBox.warning(self, "提示", f"请先上传并加载源表格{src_idx + 1}。")
            return
        if not self.target_headers:
            QMessageBox.warning(self, "提示", "请先上传并加载目标表格。")
            return

        # ---- 第一阶段：模糊匹配标准字段 ----
        fuzzy_matches = {}   # {标准字段: (源表头, 目标表头)}
        used_src = set()
        used_tgt = set()

        for std_field in self._STANDARD_FIELDS:
            src_match, src_score = self._fuzzy_find_match(std_field, src_headers)
            tgt_match, tgt_score = self._fuzzy_find_match(std_field, self.target_headers)
            if (src_match and tgt_match
                    and src_match not in used_src and tgt_match not in used_tgt):
                fuzzy_matches[std_field] = (src_match, tgt_match)
                used_src.add(src_match)
                used_tgt.add(tgt_match)

        # ---- 第二阶段：精确匹配其余相同表头 ----
        exact_matches = []
        for h in src_headers:
            if h in used_src:
                continue
            if h in self.target_headers and h not in used_tgt:
                exact_matches.append(h)
                used_src.add(h)
                used_tgt.add(h)

        # 汇总所有匹配
        all_mappings = []
        for std_field, (src_h, tgt_h) in fuzzy_matches.items():
            all_mappings.append((src_h, tgt_h, f"模糊匹配「{std_field}」"))
        for h in exact_matches:
            all_mappings.append((h, h, "精确匹配"))

        if not all_mappings:
            QMessageBox.information(
                self, "提示",
                f"源表格{src_idx + 1}和目标表没有可匹配的表头。"
            )
            return

        # 追加到映射表末尾（不清除已有行）
        start_row = self.map_table.rowCount()
        for i, (src_h, tgt_h, _) in enumerate(all_mappings):
            row = start_row + i
            self.map_table.insertRow(row)
            self.add_mapping_row(row, config={
                "method": "直接复制",
                "src_tbl_idx": src_idx,
                "src": src_h,
                "tgt": tgt_h,
                "sep": "-",
                "split": 1,
            })

        # 构建结果消息
        msg_lines = [f"已为源表格{src_idx + 1}自动匹配 {len(all_mappings)} 条表头映射："]
        for src_h, tgt_h, method_desc in all_mappings:
            msg_lines.append(f"  · {src_h}  →  {tgt_h}（{method_desc}）")

        unmatched_src = [h for h in src_headers if h not in used_src]
        if unmatched_src:
            msg_lines.append(f"\n以下源列未匹配（需手动配置）：")
            msg_lines.append("  " + ", ".join(unmatched_src))

        QMessageBox.information(self, "匹配完成", "\n".join(msg_lines))
        self.statusBar().showMessage(
            f"已为源表格{src_idx + 1}自动匹配 {len(all_mappings)} 条表头映射"
        )

    def _auto_match_headers(self):
        """旧接口兼容：对主源表（索引0）执行一键匹配。"""
        self._auto_match_headers_for(0)

    # ========== 子表选择 ==========
    def _populate_source_sheets(self):
        """加载源文件后，填充子表下拉框。"""
        path = self.source_path
        if not path:
            self.combo_source_sheet.clear()
            self.combo_source_sheet.setEnabled(False)
            self.source_sheet_name = ""
            return
        ext = path.lower()
        try:
            sheet_names = []
            if ext.endswith((".xlsx", ".xlsm")):
                wb = openpyxl.load_workbook(path, read_only=True)
                sheet_names = wb.sheetnames
                wb.close()
            elif ext.endswith(".xls"):
                try:
                    import xlrd
                    wb = xlrd.open_workbook(path)
                    sheet_names = wb.sheet_names()
                except ImportError:
                    pass
            if not sheet_names:
                sheet_names = ["Sheet1"]
        except Exception:
            sheet_names = ["Sheet1"]

        self.combo_source_sheet.blockSignals(True)
        self.combo_source_sheet.clear()
        self.combo_source_sheet.addItems(sheet_names)
        self.combo_source_sheet.setEnabled(len(sheet_names) > 1)
        self.source_sheet_name = sheet_names[0]
        self.combo_source_sheet.setCurrentIndex(0)
        self.combo_source_sheet.blockSignals(False)

    def _populate_target_sheets(self):
        """加载目标文件后，填充子表下拉框。"""
        path = self.target_path
        if not path:
            self.combo_target_sheet.clear()
            self.combo_target_sheet.setEnabled(False)
            self.target_sheet_name = ""
            return
        ext = path.lower()
        try:
            sheet_names = []
            if ext.endswith((".xlsx", ".xlsm")):
                wb = openpyxl.load_workbook(path, read_only=True)
                sheet_names = wb.sheetnames
                wb.close()
            elif ext.endswith(".xls"):
                try:
                    import xlrd
                    wb = xlrd.open_workbook(path)
                    sheet_names = wb.sheet_names()
                except ImportError:
                    pass
            if not sheet_names:
                sheet_names = ["Sheet1"]
        except Exception:
            sheet_names = ["Sheet1"]

        self.combo_target_sheet.blockSignals(True)
        self.combo_target_sheet.clear()
        self.combo_target_sheet.addItems(sheet_names)
        self.combo_target_sheet.setEnabled(len(sheet_names) > 1)
        self.target_sheet_name = sheet_names[0]
        self.combo_target_sheet.setCurrentIndex(0)
        self.combo_target_sheet.blockSignals(False)

    def _on_source_sheet_changed(self, index):
        """源表子表切换时重新读取数据。"""
        if index < 0 or not self.source_path:
            return
        sheet_name = self.combo_source_sheet.currentText()
        if sheet_name == self.source_sheet_name:
            return
        self.source_sheet_name = sheet_name
        self._apply_source_load()

    def _on_target_sheet_changed(self, index):
        """目标表子表切换时重新读取表头。"""
        if index < 0 or not self.target_path:
            return
        sheet_name = self.combo_target_sheet.currentText()
        if sheet_name == self.target_sheet_name:
            return
        self.target_sheet_name = sheet_name
        self._apply_target_load()

    def _clean_header(self, h, idx):
        if h is None or h == "":
            return f"列{idx + 1}"
        return str(h)

    def _merged_row_values(self, ws, row_1based, ncols):
        """读取指定行的单元格值；若该单元格为空但属于某合并区域，则填充合并区域左上角的值。

        这样当表头行存在跨列合并（如 A1:C1 合并为『基本信息』）时，
        被合并覆盖的列也能拿到正确的表头名，而不会出现空列。
        """
        vals = []
        for c in range(1, ncols + 1):
            v = ws.cell(row=row_1based, column=c).value
            if v is None or v == "":
                for rng in ws.merged_cells.ranges:
                    if (rng.min_row <= row_1based <= rng.max_row
                            and rng.min_col <= c <= rng.max_col):
                        v = ws.cell(row=rng.min_row, column=rng.min_col).value
                        break
            vals.append(v)
        return vals

    def read_xls(self, path, header_row=1, sheet_name=None):
        """用 xlrd 读取 .xls（按指定表头行读取表头+数据，日期以浮点返回）。"""
        try:
            import xlrd
        except ImportError:
            raise ImportError("读取 .xls 需要 xlrd，请运行: pip install xlrd")
        wb = xlrd.open_workbook(path)
        if sheet_name and sheet_name in wb.sheet_names():
            ws = wb.sheet_by_name(sheet_name)
        else:
            ws = wb.sheet_by_index(0)
        # 合并区域映射：(行,列)->左上角值
        merge_val = {}
        for (rlo, rhi, clo, chi) in ws.merged_cells:
            top = ws.cell_value(rlo, clo)
            for r in range(rlo, rhi):
                for c in range(clo, chi):
                    merge_val[(r, c)] = top
        ncols = ws.ncols
        hr = header_row - 1   # xlrd 行号 0-based
        raw = []
        for c in range(ncols):
            v = ws.cell_value(hr, c)
            if (v is None or v == "") and (hr, c) in merge_val:
                v = merge_val[(hr, c)]
            raw.append(v)
        headers = [self._clean_header(v if v != "" else None, c) for c, v in enumerate(raw)]
        data = []
        for r in range(header_row, ws.nrows):   # 数据从表头行的下一行开始
            data.append([ws.cell_value(r, c) for c in range(ncols)])
        # 表头上方内容（可能是表名/合并标题行）
        above = []
        for r in range(header_row - 1):
            above.append([ws.cell_value(r, c) for c in range(ncols)])
        return headers, data, above

    def load_source_file(self):
        path, _ = QFileDialog.getOpenFileName(
            self, "选择源表格", "",
            "Excel 文件 (*.xlsx *.xls *.xlsm)"
        )
        if not path:
            return
        self.source_path = path
        self.lbl_source.setText(f"源文件：{os.path.basename(path)}")
        self._populate_source_sheets()
        self._apply_source_load()

    def _apply_source_load(self):
        """按当前『源表表头行』读取表头与数据；表头行变化时也会触发。"""
        if not getattr(self, "source_path", None):
            return
        path = self.source_path
        hr = self.spin_source_header.value()   # 1-based 表头行
        ext = path.lower()
        try:
            if ext.endswith((".xlsx", ".xlsm")):
                wb = openpyxl.load_workbook(path, data_only=True)
                ws = wb[self.source_sheet_name] if self.source_sheet_name else wb.active
                max_row = ws.max_row
                max_col = ws.max_column
                if hr < 1 or hr > max_row:
                    QMessageBox.warning(
                        self, "表头行错误",
                        f"源表表头行 {hr} 超出范围（共 {max_row} 行）。"
                    )
                    wb.close()
                    return
                raw = self._merged_row_values(ws, hr, max_col)
                headers = [self._clean_header(c, i) for i, c in enumerate(raw)]
                data = [
                    list(r)
                    for r in ws.iter_rows(min_row=hr + 1, values_only=True)
                ]
                wb.close()
            elif ext.endswith(".xls"):
                headers, data, _ = self.read_xls(path, hr, self.source_sheet_name)
            else:
                QMessageBox.warning(self, "格式错误", "不支持的文件格式。")
                return
        except ImportError as e:
            QMessageBox.warning(self, "缺少依赖", str(e))
            return
        except Exception as e:
            QMessageBox.critical(self, "读取失败", f"无法读取源文件：{e}")
            return

        self.source_header_row = hr
        self.source_headers = headers
        self.source_data = data
        self.statusBar().showMessage(
            f"已加载源表：{len(headers)}列，{len(data)}行数据（表头第 {hr} 行）"
        )
        self.refresh_after_load()

    def load_target_file(self):
        path, _ = QFileDialog.getOpenFileName(
            self, "选择目标表格", "",
            "Excel 文件 (*.xlsx *.xls *.xlsm)"
        )
        if not path:
            return
        self.target_path = path
        self.lbl_target.setText(f"目标文件：{os.path.basename(path)}")
        self._populate_target_sheets()
        self._apply_target_load()

    def _apply_target_load(self):
        """按当前『目标表表头行』读取表头（不读数据）；表头行变化时也会触发。"""
        if not getattr(self, "target_path", None):
            return
        path = self.target_path
        hr = self.spin_target_header.value()   # 1-based 表头行
        ext = path.lower()
        try:
            if ext.endswith((".xlsx", ".xlsm")):
                wb = openpyxl.load_workbook(path)  # 保留格式/公式
                ws = wb[self.target_sheet_name] if self.target_sheet_name else wb.active
                max_row = ws.max_row
                if hr < 1 or hr > max_row:
                    QMessageBox.warning(
                        self, "表头行错误",
                        f"目标表表头行 {hr} 超出范围（共 {max_row} 行）。"
                    )
                    wb.close()
                    return
                raw = self._merged_row_values(ws, hr, ws.max_column)
                self.target_headers = [self._clean_header(h, c) for c, h in enumerate(raw)]
                self.statusBar().showMessage(
                    f"目标表表头共 {len(self.target_headers)} 列（表头第 {hr} 行）"
                )
                wb.close()
            elif ext.endswith(".xls"):
                headers, _, _ = self.read_xls(path, hr, self.target_sheet_name)
                self.target_headers = [self._clean_header(h, c) for c, h in enumerate(headers)]
                QMessageBox.information(
                    self, "提示",
                    ".xls 目标表仅读取表头，输出将生成新的 .xlsx 文件，\n"
                    "原文件的格式/样式/公式不会被保留。"
                )
                self.statusBar().showMessage(
                    f"目标表表头共 {len(self.target_headers)} 列（.xls，表头第 {hr} 行）"
                )
            else:
                QMessageBox.warning(self, "格式错误", "不支持的文件格式。")
                return
        except ImportError as e:
            QMessageBox.warning(self, "缺少依赖", str(e))
            return
        except Exception as e:
            QMessageBox.critical(self, "读取失败", f"无法读取目标文件：{e}")
            return

        self.target_header_row = hr
        self.target_col_count = len(self.target_headers)
        self.refresh_after_load()

    # ========== 映射预览 ==========
    def preview_mapping(self):
        """预览数据映射关系：显示每条规则的前5行数据样例。"""
        loaded_sources = [i for i, s in enumerate(self.sources) if s.get("path") and s.get("headers")]
        if not loaded_sources:
            QMessageBox.warning(self, "提示", "请先上传并加载至少一个源表格。")
            return
        if not self.target_headers:
            QMessageBox.warning(self, "提示", "请先上传并加载目标表格。")
            return

        # 收集已配置的规则
        rules = []
        for row in range(self.map_table.rowCount()):
            method = self.map_table.cellWidget(row, COL_METHOD).currentText()
            src_sel = self.get_source_selection(row)
            tgt_sel = self.get_target_selection(row)
            src_idx = self._get_source_idx_for_row(row)

            if src_idx >= len(self.sources) or not self.sources[src_idx].get("headers"):
                continue

            if method == "合并复制":
                if not (isinstance(src_sel, list) and len(src_sel) >= 2) or not tgt_sel:
                    continue
                sep_w = self.map_table.cellWidget(row, COL_SEP)
                sep_text = sep_w.text() if sep_w else "-"
                rules.append(("merge", src_idx, src_sel, tgt_sel, sep_text))
            elif method == "拆分复制":
                n = self.map_table.cellWidget(row, COL_SPLIT).value()
                if (not src_sel
                        or not (isinstance(tgt_sel, list) and len(tgt_sel) == n)
                        or len(set(tgt_sel)) != len(tgt_sel)):
                    continue
                rules.append(("split", src_idx, src_sel, tgt_sel, n))
            else:
                if not src_sel or not tgt_sel:
                    continue
                rules.append(("direct", src_idx, src_sel, tgt_sel, None))

        if not rules:
            QMessageBox.information(self, "预览", "没有已配置的映射规则可预览。")
            return

        # 构建预览表格
        preview_rows = []
        for rule in rules:
            kind = rule[0]
            src_idx = rule[1]
            src_info = self.sources[src_idx]
            src_headers = src_info.get("headers", [])
            src_data = src_info.get("data", [])
            sample_count = min(5, len(src_data))

            if kind == "direct":
                _, _, src_name, tgt_name, _ = rule
                if src_name not in src_headers or tgt_name not in self.target_headers:
                    preview_rows.append((f"源表{src_idx+1}→{tgt_name}", [src_name], ["(列名不匹配)"]))
                    continue
                scol = src_headers.index(src_name)
                samples = []
                for i in range(sample_count):
                    val = src_data[i][scol] if scol < len(src_data[i]) else None
                    samples.append(str(normalize(val)) if val is not None else "")
                preview_rows.append((f"源表{src_idx+1}·{src_name} → {tgt_name}", [src_name], samples))
            elif kind == "merge":
                _, _, src_list, tgt_name, sep = rule
                scols = []
                ok = True
                for s in src_list:
                    if s not in src_headers:
                        ok = False
                        break
                    scols.append(src_headers.index(s))
                if not ok:
                    preview_rows.append((f"源表{src_idx+1}→{tgt_name}(合并)", src_list, ["(列名不匹配)"]))
                    continue
                samples = []
                for i in range(sample_count):
                    parts = []
                    for cc in scols:
                        v = src_data[i][cc] if cc < len(src_data[i]) else None
                        parts.append(str(normalize(v)) if v is not None else "")
                    samples.append(sep.join(parts))
                preview_rows.append((f"源表{src_idx+1}·{'+'.join(src_list)} → {tgt_name}(合并)", src_list, samples))
            elif kind == "split":
                _, _, src_name, tgt_cols, nn = rule
                if src_name not in src_headers:
                    preview_rows.append((f"源表{src_idx+1}→{','.join(tgt_cols)}(拆分)", [src_name], ["(列名不匹配)"]))
                    continue
                scol = src_headers.index(src_name)
                samples = []
                for i in range(sample_count):
                    val = src_data[i][scol] if scol < len(src_data[i]) else None
                    s = str(normalize(val)) if val is not None else ""
                    samples.append(s)
                preview_rows.append((f"源表{src_idx+1}·{src_name} → {','.join(tgt_cols)}(拆分)", [src_name], samples))

        # 弹出预览对话框
        dlg = QDialog(self)
        dlg.setWindowTitle("映射预览（前5行数据样例）")
        dlg.setMinimumSize(800, 500)
        layout = QVBoxLayout(dlg)

        info = QLabel(f"共 {len(rules)} 条映射规则，以下为每条规则的前 5 行数据预览：")
        layout.addWidget(info)

        table = QTableWidget()
        table.setColumnCount(7)
        table.setHorizontalHeaderLabels(["序号", "映射关系", "源表", "源列", "目标列", "方式", "数据样例(前5行)"])
        table.setRowCount(len(preview_rows))
        table.horizontalHeader().setSectionResizeMode(QHeaderView.Stretch)
        table.setEditTriggers(QAbstractItemView.NoEditTriggers)

        for i, (desc, src_cols, samples) in enumerate(preview_rows):
            # 解析描述获取各字段
            parts = desc.split("·", 1)
            src_tbl = parts[0] if len(parts) > 1 else ""
            rest = parts[1] if len(parts) > 1 else desc
            if " → " in rest:
                src_col_part, tgt_part = rest.split(" → ", 1)
            else:
                src_col_part, tgt_part = rest, ""
            method_tag = ""
            if "(合并)" in tgt_part:
                method_tag = "合并复制"
                tgt_part = tgt_part.replace("(合并)", "")
            elif "(拆分)" in tgt_part:
                method_tag = "拆分复制"
                tgt_part = tgt_part.replace("(拆分)", "")
            else:
                method_tag = "直接复制"

            table.setItem(i, 0, QTableWidgetItem(str(i + 1)))
            table.setItem(i, 1, QTableWidgetItem(desc))
            table.setItem(i, 2, QTableWidgetItem(src_tbl))
            table.setItem(i, 3, QTableWidgetItem(", ".join(src_cols)))
            table.setItem(i, 4, QTableWidgetItem(tgt_part))
            table.setItem(i, 5, QTableWidgetItem(method_tag))
            table.setItem(i, 6, QTableWidgetItem(" | ".join(samples)))

        layout.addWidget(table, 1)

        btn_close = QPushButton("关闭")
        btn_close.clicked.connect(dlg.accept)
        layout.addWidget(btn_close)
        dlg.exec_()

    # ========== 执行与保存 ==========
    def do_copy(self):
        # 检查至少有一个已加载的源表
        loaded_sources = [i for i, s in enumerate(self.sources) if s.get("path") and s.get("headers")]
        if not loaded_sources:
            QMessageBox.warning(self, "警告", "请先上传并加载至少一个源表格。")
            return
        if not self.target_headers:
            QMessageBox.warning(self, "警告", "请先上传并加载目标表格。")
            return

        # 仅处理「已配置」的行；未配置的行静默跳过。
        # 每条规则额外携带 src_idx（源表索引），用于多源数据读取。
        rules = []
        for row in range(self.map_table.rowCount()):
            method = self.map_table.cellWidget(row, COL_METHOD).currentText()
            src_sel = self.get_source_selection(row)
            tgt_sel = self.get_target_selection(row)
            src_idx = self._get_source_idx_for_row(row)

            # 跳过未加载的源表
            if src_idx >= len(self.sources) or not self.sources[src_idx].get("headers"):
                continue

            if method == "合并复制":
                if not (isinstance(src_sel, list) and len(src_sel) >= 2) or not tgt_sel:
                    continue
                sep = self.map_table.cellWidget(row, COL_SEP)
                sep_text = sep.text() if sep else "-"
                rules.append(("merge", src_idx, src_sel, tgt_sel, sep_text))
            elif method == "拆分复制":
                n = self.map_table.cellWidget(row, COL_SPLIT).value()
                if (not src_sel
                        or not (isinstance(tgt_sel, list) and len(tgt_sel) == n)
                        or len(set(tgt_sel)) != len(tgt_sel)):
                    continue
                rules.append(("split", src_idx, src_sel, tgt_sel, n))
            else:  # 直接复制
                if not src_sel or not tgt_sel:
                    continue
                rules.append(("direct", src_idx, src_sel, tgt_sel, None))

        if not rules:
            QMessageBox.warning(
                self, "提示",
                "没有已配置的映射规则。\n请至少为一行选择「源列」与「目标列」后再复制。"
            )
            return

        # 选择输出路径
        default_dir = os.path.dirname(self.target_path or self.source_path or ".")
        target_base = os.path.splitext(os.path.basename(self.target_path))[0] if self.target_path else "output"
        ts = datetime.datetime.now().strftime("%Y%m%d_%H%M")
        default_name = f"{target_base}_{ts}.xlsx"
        out, _ = QFileDialog.getSaveFileName(
            self, "另存为", os.path.join(default_dir, default_name),
            "Excel 文件 (*.xlsx)"
        )
        if not out:
            return
        if not out.lower().endswith(".xlsx"):
            out += ".xlsx"

        try:
            self.write_output(rules, out)
            QMessageBox.information(
                self, "完成",
                f"已复制 {len(rules)} 条映射规则，处理完成。\n已保存至：\n{out}"
            )
            self.statusBar().showMessage(f"已生成文件：{os.path.basename(out)}")
        except Exception as e:
            QMessageBox.critical(self, "错误", f"保存失败：{e}")

    def write_output(self, rules, out):
        # 加载目标工作簿（.xlsx 保留格式；.xls 则基于表头新建）
        data_start = self.target_header_row + 1   # 数据从表头行的下一行开始
        if self.target_path and self.target_path.lower().endswith((".xlsx", ".xlsm")):
            wb = openpyxl.load_workbook(self.target_path)
            ws = wb[self.target_sheet_name] if self.target_sheet_name else wb.active
        else:
            wb = openpyxl.Workbook()
            ws = wb.active
            for c, h in enumerate(self.target_headers):
                ws.cell(row=self.target_header_row, column=c + 1, value=h)

        def safe_write(ws, row, col, value):
            """安全写入单元格：若该单元格属于合并区域（非左上角），先解除合并再写入。"""
            cell = ws.cell(row=row, column=col)
            if hasattr(cell, 'value') and not isinstance(cell, openpyxl.cell.cell.MergedCell):
                cell.value = value
                return
            for rng in list(ws.merged_cells.ranges):
                if rng.min_row <= row <= rng.max_row and rng.min_col <= col <= rng.max_col:
                    ws.unmerge_cells(str(rng))
                    break
            ws.cell(row=row, column=col, value=value)

        # 多源数据复制：每条规则包含 src_idx，从对应源表读取数据
        # 不同源表的行数可能不同，各自独立写入
        for rule in rules:
            kind = rule[0]
            src_idx = rule[1]

            # 获取该规则对应源表的表头和数据
            src_info = self.sources[src_idx] if src_idx < len(self.sources) else {}
            src_headers = src_info.get("headers", [])
            src_data = src_info.get("data", [])
            n = len(src_data)

            if kind == "direct":
                _, _, src_name, tgt_name, _ = rule
                if src_name not in src_headers or tgt_name not in self.target_headers:
                    continue  # 列名不匹配，跳过
                scol = src_headers.index(src_name)
                tcol = self.target_headers.index(tgt_name)
                for i in range(n):
                    rowdata = src_data[i]
                    val = rowdata[scol] if scol < len(rowdata) else None
                    safe_write(ws, data_start + i, tcol + 1, normalize(val))
            elif kind == "merge":
                _, _, src_list, tgt_name, sep = rule
                if tgt_name not in self.target_headers:
                    continue
                # 检查所有源列是否存在
                missing = [s for s in src_list if s not in src_headers]
                if missing:
                    continue
                scols = [src_headers.index(s) for s in src_list]
                tcol = self.target_headers.index(tgt_name)
                for i in range(n):
                    rowdata = src_data[i]
                    parts = []
                    for cc in scols:
                        v = rowdata[cc] if cc < len(rowdata) else None
                        parts.append(str(normalize(v)) if v is not None else "")
                    safe_write(ws, data_start + i, tcol + 1, sep.join(parts))
            elif kind == "split":
                _, _, src_name, tgt_cols, nn = rule
                if src_name not in src_headers:
                    continue
                scol = src_headers.index(src_name)
                for i in range(n):
                    rowdata = src_data[i]
                    v = rowdata[scol] if scol < len(rowdata) else None
                    s = str(normalize(v)) if v is not None else ""
                    for tname in tgt_cols:
                        if tname in self.target_headers:
                            tcol = self.target_headers.index(tname)
                            safe_write(ws, data_start + i, tcol + 1, s)
        wb.save(out)


# ---------- 入口 ----------
def main():
    app = QApplication(sys.argv)
    # 麒麟 Linux 常见中文字体，缺失时自动回退
    for font_name in ("WenQuanYi Micro Hei", "Noto Sans CJK SC", "Source Han Sans SC"):
        f = QFont(font_name, 10)
        if f.exactMatch() or font_name == "WenQuanYi Micro Hei":
            app.setFont(f)
            break
    win = MainWindow()
    win.showMaximized()   # 启动时全屏最大化
    sys.exit(_app_exec(app))


if __name__ == "__main__":
    main()
